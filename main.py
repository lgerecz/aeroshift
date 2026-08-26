import re
import traceback
import os
import json
import base64
import sys
import io
import contextlib
import time
import unicodedata
from fastapi import FastAPI, HTTPException, Header, Body, UploadFile, File, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Dict, Any, Optional
from ortools.sat.python import cp_model

app = FastAPI(
    title="AeroShift Backend",
    description="Backend de Optimización (OR-Tools) y Visión IA para AeroShift - Versión Azul Handling",
    version="1.0.0"
)

# Enable CORS for frontend communication
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Permits all origins for easy development/testing
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 1. ZONAS (FROM YOUR COLAB CODE)
SCHENGEN = {
    'FRA','MUC','BER','DUS','HAM','STR','CGN','NUE','HAJ','BRE','FMO','NRN',
    'CDG','ORY','LYS','NCE','MRS','TLS','BOD','NTE','LIL','BVA',
    'FCO','MXP','LIN','VCE','NAP','BGY','PSA','CIA','TRN','BLQ','PMO','CAG','BRI','TSF',
    'AMS','EIN','RTM','BRU','CRL','LIS','OPO','FAO',
    'ATH','SKG','HER','RHO','CFU','JMK','ZTH','KGS','CHQ',
    'ARN','GOT','MMX','VST','OSL','BGO','SVG','TRD',
    'CPH','BLL','AAL','AAR','HEL','TMP','TKU','OUL',
    'VIE','GRZ','INN','SZG','LNZ','PRG','BRQ','OSR',
    'WAW','KRK','GDN','WRO','KTW','POZ','LCJ','RZE','WMI',
    'BUD','OTP','CLJ','TSR','IAS','BTS','LJU',
    'ZAG','SPU','DBV','PUL','ZAD','RJK','SOF','VAR','BOJ',
    'LUX','ZRH','GVA','BSL','KEF','MLA','TLL','RIX','VNO',
    'MAD','BCN','PMI','VLC','SVQ','AGP','BIO','SDR','ZAZ',
    'TFN','TFS','LPA','ACE','FUE','VDE','SPC','GMZ','MLN','IBZ','MAH',
}

NON_SCHENGEN = {
    'LHR','LGW','LCY','STN','LTN','MAN','BHX','EDI','GLA',
    'BRS','NCL','LPL','EMA','SOU','ABZ','BFS','PIK','LBA','CWL','BHD','MME','BOH',
    'DUB','ORK','SNN','NOC','KIR',
    'CMN','RAK','AGA','TNG','FEZ','RBA','OZZ','NDR','SFI',
    'IST','SAW','AYT','ADB','ESB','GZT','DLM','BJV',
    'TUN','MIR','SFA','DJE','CAI','HRG','SSH','LXR',
    'ALG','ORN','CZL','AAE','TLM','BJA','GJL',
    'DKR','SID','RAI','BVC','VXE',
    'TLV','DXB','AUH','SHJ','DOH','KWI','RUH','JED',
    'JFK','LAX','MIA','ORD','ATL','DFW','SFO','BOS',
    'MEX','CUN','SDQ','PUJ','HAV','BOG','MDE','CCS','LIM','GRU','GIG','EZE','SCL',
    'BEG','TIA','PRN','TGD','TIV','SJJ','TBS','EVN','KIV',
}

def get_zona(iata: str) -> str:
    c = iata.upper().strip()
    if c in SCHENGEN:
        return 'S'
    return 'NS'

def gap_minimo(za: str, zb: str) -> int:
    return 60 if za == zb else 75

def gap_tolerancia(za: str, zb: str) -> int:
    return 55 if za == zb else 70

PUERTAS_REMOTAS = {'B20','B22','C32','C36','C39','C40'}
ROLES_NO_EMBARCAN = {'DSM','PSM','OPS','TKT','TKD','LL','SOMBRA','SHADOW','SHADOWING','FAMI','SICK','CURSO','NUEVO','NEW','AUTOCHECKIN'}
ROLES_OPERATIVOS = {'TKT','LL','OPS'}


def get_base_role(role: str) -> str:
    """Devuelve el rol estructural aunque exista un estado: TKT (SICK) -> TKT."""
    normalized = str(role or "").upper().strip()
    match = re.match(r"^([A-ZÁÉÍÓÚÜÑ]+)", normalized)
    return match.group(1) if match else ""


def get_full_shift_status(role: str) -> Optional[str]:
    """Detecta estados sin intervalo que anulan toda la jornada."""
    normalized = str(role or "").upper().strip()
    if re.search(r"\d{1,2}:\d{2}", normalized):
        return None
    match = re.search(r"\((SICK|NUEVO|NEW)\)", normalized)
    if match:
        return "NUEVO" if match.group(1) == "NEW" else match.group(1)
    return None


def is_non_boarding_role(role: str) -> bool:
    """Reconoce roles base, estados completos y motivos compuestos."""
    normalized = str(role or "").upper().strip()
    if normalized in ROLES_NO_EMBARCAN:
        return True
    if get_base_role(normalized) in ROLES_NO_EMBARCAN:
        return True
    return get_full_shift_status(normalized) in {"SICK", "NUEVO"}

# Helper time functions
def hms(t: str) -> int:
    try:
        h, m = map(int, t.split(':'))
        return h * 60 + m
    except Exception:
        return 0


def normalize_and_measure_schedule(schedule: str) -> tuple[str, Optional[int]]:
    """
    Normaliza uno o dos tramos y calcula los minutos realmente trabajados.
    Si la salida es anterior a la entrada, se interpreta como el día siguiente.
    Devuelve (horario_normalizado, minutos) o ("ILEGIBLE", None).
    """
    text = str(schedule or "").strip().upper()
    if text == "ILEGIBLE":
        return "ILEGIBLE", None

    text = text.replace("–", "-").replace("—", "-").replace("//", "/")
    raw_segments = [segment.strip() for segment in text.split("/") if segment.strip()]
    if not 1 <= len(raw_segments) <= 2:
        raise ValueError(f"Número de tramos inválido: {schedule}")

    normalized_segments = []
    total_minutes = 0
    pattern = re.compile(r"^(\d{1,2}):(\d{2})\s*-\s*(\d{1,2}):(\d{2})$")

    for segment in raw_segments:
        match = pattern.fullmatch(segment)
        if not match:
            raise ValueError(f"Formato de horario inválido: {schedule}")
        start_h, start_m, end_h, end_m = map(int, match.groups())
        if start_h > 23 or end_h > 23 or start_m > 59 or end_m > 59:
            raise ValueError(f"Hora fuera de rango: {schedule}")

        start = start_h * 60 + start_m
        end = end_h * 60 + end_m
        if end < start:
            end += 1440
        duration = end - start
        if duration <= 0:
            raise ValueError(f"Tramo de duración nula: {schedule}")

        total_minutes += duration
        normalized_segments.append(
            f"{start_h:02d}:{start_m:02d}-{end_h:02d}:{end_m:02d}"
        )

    return " / ".join(normalized_segments), total_minutes


def _normalize_clock_token(token: str) -> str:
    """Normaliza 17, 7:30 o 17:00 al formato HH:MM."""
    value = str(token or "").strip()
    match = re.fullmatch(r"(\d{1,2})(?::(\d{2}))?", value)
    if not match:
        raise ValueError(f"Hora de restricción inválida: {token}")
    hour = int(match.group(1))
    minute = int(match.group(2) or "0")
    if hour > 23 or minute > 59:
        raise ValueError(f"Hora de restricción fuera de rango: {token}")
    return f"{hour:02d}:{minute:02d}"


def normalize_mixed_role(role: str) -> tuple[str, Optional[tuple[str, str, str]]]:
    """
    Normaliza restricciones mixtas al formato que consumen OR-Tools y el
    optimizador local: CSA (HH:MM-HH:MM ROL).

    Admite, entre otras, estas variantes generadas a partir del documento:
    - CSA (12:30-17:00 OPS)
    - CSA (OPS 12:30-17)
    - OPS (12:30-17)
    - OPS 12:30-17
    """
    text = str(role or "").upper().strip().replace("TKD", "TKT")
    time_token = r"\d{1,2}(?::\d{2})?"
    role_token = r"[A-ZÁÉÍÓÚÜÑ]+(?:\s+[A-ZÁÉÍÓÚÜÑ]+)*"

    patterns = [
        # Hora-hora seguida del rol restringido.
        rf"^(?:CSA\s*)?\(\s*({time_token})\s*-\s*({time_token})\s+({role_token})\s*\)$",
        # Rol restringido seguido de hora-hora, dentro de paréntesis.
        rf"^(?:CSA\s*)?\(\s*({role_token})\s+({time_token})\s*-\s*({time_token})\s*\)$",
        # Rol seguido del intervalo, con paréntesis solo alrededor de las horas.
        rf"^({role_token})\s*\(\s*({time_token})\s*-\s*({time_token})\s*\)$",
        # Rol e intervalo sin paréntesis.
        rf"^({role_token})\s+({time_token})\s*-\s*({time_token})$",
    ]

    for pattern_index, pattern in enumerate(patterns):
        match = re.fullmatch(pattern, text)
        if not match:
            continue
        groups = match.groups()
        if pattern_index == 0:
            start_raw, end_raw, restricted_role = groups
        else:
            restricted_role, start_raw, end_raw = groups
        start = _normalize_clock_token(start_raw)
        end = _normalize_clock_token(end_raw)
        restricted_role = restricted_role.upper().replace("TKD", "TKT")
        canonical = f"CSA ({start}-{end} {restricted_role})"
        return canonical, (start, end, restricted_role)

    return text, None


def mixed_interval_is_within_schedule(
    schedule: str, start: str, end: str
) -> bool:
    """Comprueba que la restricción mixta cae dentro de algún tramo trabajado."""
    if schedule == "ILEGIBLE":
        return True

    def to_minutes(clock: str) -> int:
        hour, minute = map(int, clock.split(":"))
        return hour * 60 + minute

    work_segments = []
    previous_end = None
    for raw_segment in schedule.split(" / "):
        segment_start, segment_end = raw_segment.split("-")
        start_minutes = to_minutes(segment_start)
        end_minutes = to_minutes(segment_end)
        if end_minutes < start_minutes:
            end_minutes += 1440
        if previous_end is not None and start_minutes < previous_end:
            start_minutes += 1440
            if end_minutes < start_minutes:
                end_minutes += 1440
        work_segments.append((start_minutes, end_minutes))
        previous_end = end_minutes

    restriction_start = to_minutes(start)
    restriction_end = to_minutes(end)
    if restriction_end < restriction_start:
        restriction_end += 1440

    # Se prueban tanto el día base como el día siguiente para turnos nocturnos.
    for offset in (0, 1440):
        candidate_start = restriction_start + offset
        candidate_end = restriction_end + offset
        if any(
            candidate_start >= work_start and candidate_end <= work_end
            for work_start, work_end in work_segments
        ):
            return True
    return False


def t2m_fin(ini_s: str, fin_s: str) -> int:
    i = hms(ini_s)
    f = hms(fin_s)
    if f == 0 or f < i:
        f += 1440
    return f

def parse_mixed_role_exclusion(role_str: str) -> Optional[tuple]:
    """
    Parses a mixed role string like 'CSA (11:15-15:00 TKT)' and returns
    the interval in minutes (inicio, fin) and the restricted role, or None.
    """
    if not role_str:
        return None
    # Admite motivos simples o compuestos: TKT, OPS, CURSO TKT, SOMBRA TKT...
    match = re.search(
        r"\((\d{2}:\d{2})-(\d{2}:\d{2})\s+([A-ZÁÉÍÓÚÜÑ]+(?:\s+[A-ZÁÉÍÓÚÜÑ]+)*)\)",
        role_str.upper(),
    )
    if match:
        ini_str, fin_str, r_role = match.groups()
        return hms(ini_str), t2m_fin(ini_str, fin_str), r_role.upper().strip()
    return None

def m2t(m: int) -> str:
    m = m % 1440
    return f"{m//60:02d}:{m%60:02d}"

def dur_str(m: int) -> str:
    return f"{m//60}h {m%60:02d}min"

def hor_turno_completo(ag: dict) -> str:
    txt = f"{ag['inicio']}–{ag['fin']}"
    if ag.get('bloque2'):
        txt += f" / {ag['bloque2']['inicio']}–{ag['bloque2']['fin']}"
    txt += f" / {dur_str(ag['_jornada'])}"
    return txt

def ndisp(ag: dict) -> str:
    if get_base_role(ag['rol']) in ROLES_OPERATIVOS:
        return f"{ag['nombre']} [{ag['rol']}]"
    if ag['espec']:
        return f"{ag['nombre']} [{'/'.join(ag['espec'])}]"
    return ag['nombre']

def icono_break(mins: int) -> str:
    if mins >= 70:
        return "✅"
    if mins >= 55:
        return "⏳"
    return ""

def vkey(v: dict) -> str:
    return f"{v['codigo']}_{v['num']}"

# Pydantic Schemas for API
class AgentInput(BaseModel):
    id: int
    name: str
    inicio: str
    fin: str
    rol: str
    espec: List[str] = []
    excluir: bool = False
    excluir_embarque: bool = False
    bloque2: Optional[Dict[str, str]] = None

class FlightInput(BaseModel):
    id: int
    destination: str
    number: str
    time: str              # departure time (STD)
    gate: Optional[str] = ""
    pax: Optional[int] = None
    charter: Optional[bool] = False
    manual: Optional[bool] = False

class OptimizeRequest(BaseModel):
    agents: List[AgentInput]
    flights: List[FlightInput]
    modo: Optional[str] = 'PROPORCIONAL' # 'EQUILIBRADO' | 'PROPORCIONAL'

@app.post("/optimize")
def optimize_schedule(req: OptimizeRequest):
    """
    Usa Google OR-Tools (CP-SAT Solver) para resolver la asignación óptima de agentes a vuelos
    siguiendo estrictamente el modelo matemático y generando los 5 reportes de tu script de Google Colab.
    """
    agents_input = req.agents
    flights_input = req.flights
    MODO = req.modo or 'PROPORCIONAL'
    
    if not agents_input:
        raise HTTPException(status_code=400, detail="No hay agentes disponibles para optimizar.")
    if not flights_input:
        raise HTTPException(status_code=400, detail="No hay vuelos para asignar.")
    incomplete_flights = [
        flight.id for flight in flights_input
        if not str(flight.destination or "").strip()
        or not str(flight.number or "").strip()
        or not str(flight.time or "").strip()
        or flight.pax is None
    ]
    if incomplete_flights:
        raise HTTPException(
            status_code=400,
            detail=(
                "Hay vuelos con datos pendientes de revisión: "
                + ", ".join(map(str, incomplete_flights))
            ),
        )

    # 1. PREPROCESAMIENTO (FROM COLAB SECTION 6)
    AGENTES = []
    for ag in agents_input:
        ag_dict = {
            'id': ag.id,
            'nombre': ag.name,
            'inicio': ag.inicio,
            'fin': ag.fin,
            'rol': ag.rol,
            'espec': ag.espec,
            'excluir': ag.excluir,
            'excluir_embarque': ag.excluir_embarque,
            'bloque2': ag.bloque2
        }
        # Time processing
        ag_dict['_t_ini'] = hms(ag.inicio)
        if ag.bloque2:
            ag_dict['_pausa_ini'] = t2m_fin(ag.inicio, ag.fin)
            b2i = hms(ag.bloque2['inicio'])
            b2f = hms(ag.bloque2['fin'])
            if b2i < ag_dict['_pausa_ini'] % 1440:
                b2i += 1440
            if b2f < b2i:
                b2f += 1440
            ag_dict['_pausa_fin'] = b2i
            ag_dict['_t_fin'] = b2f
            ag_dict['_jornada'] = (ag_dict['_pausa_ini'] - ag_dict['_t_ini']) + (ag_dict['_t_fin'] - ag_dict['_pausa_fin'])
        else:
            ag_dict['_pausa_ini'] = ag_dict['_pausa_fin'] = None
            ag_dict['_t_fin'] = t2m_fin(ag.inicio, ag.fin)
            ag_dict['_jornada'] = ag_dict['_t_fin'] - ag_dict['_t_ini']
        
        ag_dict['_midpoint'] = ag_dict['_t_ini'] + ag_dict['_jornada'] // 2
        
        # Check for mixed role exclusions or split shifts
        ag_dict['_excl_intervals'] = []
        if ag_dict.get('bloque2'):
            ag_dict['_excl_intervals'].append((ag_dict['_pausa_ini'], ag_dict['_pausa_fin']))
        
        mixed_info = parse_mixed_role_exclusion(ag_dict['rol'])
        if mixed_info:
            ex_ini, ex_fin, restricted_role = mixed_info
            if is_non_boarding_role(restricted_role):
                ag_dict['_excl_intervals'].append((ex_ini, ex_fin))
                
        AGENTES.append(ag_dict)

    # Filter pools
    activos = [a for a in AGENTES if not a['excluir'] and not a.get('excluir_embarque') and not is_non_boarding_role(a['rol'])]
    cobertura_pool = [a for a in AGENTES if not a['excluir'] and a['espec']]
    activos_idx = {ag['id']: ai for ai, ag in enumerate(activos)}
    solo_cobertura = [a for a in AGENTES if a.get('excluir_embarque') and not a['excluir'] and not is_non_boarding_role(a['rol'])]
    todos_csa = activos + solo_cobertura

    # Process flights
    VUELOS = []
    for idx, v in enumerate(flights_input):
        std = hms(v.time)
        v_dict = {
            'num': idx + 1,
            'id': v.id,
            'codigo': v.number,
            'destino': v.destination,
            'std': v.time,
            'puerta': v.gate,
            'pax': v.pax,
            'charter': v.charter,
            'manual': v.manual,
            'emb_inicio': std - 40,
            'emb_fin': std - 15,
            'std_min': std,
            'zona': get_zona(v.destination),
            'agentes_req': 3 if v.gate in PUERTAS_REMOTAS else 2,
            'pax_unico_ok': v.pax <= 100
        }
        VUELOS.append(v_dict)

    VUELOS_ORD = sorted(VUELOS, key=lambda v: v['std_min'])
    SPLIT = 660

    # 2. MODELO CP-SAT (FROM COLAB SECTION 7)
    model = cp_model.CpModel()
    solver = cp_model.CpSolver()
    
    A, V = len(activos), len(VUELOS)
    if A == 0 or V == 0:
        raise HTTPException(status_code=400, detail="No hay suficientes agentes activos o vuelos para realizar el cálculo.")

    x = [[model.NewBoolVar(f'x_{a}_{v}') for v in range(V)] for a in range(A)]
    TOL_ENT, TOL_SAL = 10, 5

    # Constraint: Flight agents requirement
    for vi, v in enumerate(VUELOS):
        total = sum(x[ai][vi] for ai in range(A))
        if v['pax_unico_ok']:
            model.Add(total >= 1)
            model.Add(total <= v['agentes_req'])
        else:
            model.Add(total == v['agentes_req'])

    # Constraint: Availability bounds & double blocks
    for ai, ag in enumerate(activos):
        disp_desde = ag['_t_ini'] + 30 - TOL_ENT
        for vi, v in enumerate(VUELOS):
            disp_hasta = ag['_t_fin'] + TOL_SAL - 15
            if v['emb_inicio'] < disp_desde or v['std_min'] > disp_hasta:
                model.Add(x[ai][vi] == 0)
                continue
            
            # Check overlap with exclusion intervals (like split shift pause or mixed role TKT interval)
            overlap = False
            for ex_ini, ex_fin in ag.get('_excl_intervals', []):
                if v['emb_inicio'] < ex_fin and v['emb_fin'] > ex_ini:
                    overlap = True
                    break
                # Special check: flight starts before pause, but departs after pause starts + TOL_SAL
                # (applies only to the split shift pause to prevent agent getting trapped at gate)
                if ag.get('bloque2') and ex_ini == ag['_pausa_ini']:
                    if v['emb_inicio'] < ag['_pausa_ini'] and v['std_min'] > ag['_pausa_ini'] + TOL_SAL:
                        overlap = True
                        break
            if overlap:
                model.Add(x[ai][vi] == 0)

    # Constraint: No overlapping flights (simultaneous flights)
    for ai in range(A):
        for vi in range(V):
            for vj in range(vi+1, V):
                va, vb = VUELOS[vi], VUELOS[vj]
                if va['emb_inicio'] < vb['emb_fin'] and vb['emb_inicio'] < va['emb_fin']:
                    model.Add(x[ai][vi] + x[ai][vj] <= 1)

    # Constraint: Gap times (same-zone / different-zone separation)
    for ai in range(A):
        for vi in range(V):
            for vj in range(V):
                if vi == vj:
                    continue
                va, vb = VUELOS[vi], VUELOS[vj]
                if vb['emb_inicio'] <= va['emb_inicio']:
                    continue
                if vb['emb_inicio'] - va['emb_inicio'] < gap_tolerancia(va['zona'], vb['zona']):
                    model.Add(x[ai][vi] + x[ai][vj] <= 1)

    # Optional flight interval variables for breaks/coverage
    flight_iv_dict = {}
    for ai, ag in enumerate(activos):
        if ag['_jornada'] > 360 or ag['espec']:
            flight_iv_dict[ai] = [
                model.NewOptionalIntervalVar(
                    v['emb_inicio'], 
                    v['std_min'] - v['emb_inicio'], 
                    v['std_min'], 
                    x[ai][vi], 
                    f'fi_{ai}_{vi}'
                ) for vi, v in enumerate(VUELOS)
            ]

    # Constraint: Lunch Breaks
    break_iv_dict = {}
    for ai, ag in enumerate(activos):
        if ag['_jornada'] > 360:
            mid = ag['_midpoint']
            hi = max(mid, ag['_t_fin'] - 70)
            brk_s = model.NewIntVar(mid, hi, f'bs_{ai}')
            brk_e = model.NewIntVar(mid+70, ag['_t_fin'], f'be_{ai}')
            brk_iv = model.NewIntervalVar(brk_s, 70, brk_e, f'bi_{ai}')
            break_iv_dict[ai] = brk_iv
            model.AddNoOverlap([brk_iv] + flight_iv_dict[ai])

    # Constraint: Department Coverage (TKT, LL, OPS)
    for op_idx, op_ag in enumerate(AGENTES):
        base_operational_role = get_base_role(op_ag['rol'])
        if (
            base_operational_role not in ROLES_OPERATIVOS
            or op_ag['excluir']
            or get_full_shift_status(op_ag['rol']) == 'SICK'
            or op_ag['_jornada'] <= 360
        ):
            continue
        dept = base_operational_role
        op_mid = op_ag['_midpoint']
        op_fin = op_ag['_t_fin']
        if op_fin - op_mid < 70:
            continue
        
        can_cover = []
        for cov_ag in cobertura_pool:
            if dept not in cov_ag['espec']:
                continue
            if cov_ag['_t_fin'] <= op_mid or cov_ag['_t_ini'] >= op_fin:
                continue
            cov_from = max(op_mid, cov_ag['_t_ini'])
            cov_to = min(op_fin, cov_ag['_t_fin'])
            if cov_to - cov_from < 70:
                continue
            
            is_cov = model.NewBoolVar(f'cov_{cov_ag["nombre"]}_{op_idx}')
            cov_s = model.NewIntVar(cov_from, cov_to-70, f'cs_{cov_ag["nombre"]}_{op_idx}')
            cov_e = model.NewIntVar(cov_from+70, cov_to, f'ce_{cov_ag["nombre"]}_{op_idx}')
            cov_iv = model.NewOptionalIntervalVar(cov_s, 70, cov_e, is_cov, f'civ_{cov_ag["nombre"]}_{op_idx}')
            
            no_ov = [cov_iv]
            ai = activos_idx.get(cov_ag['id'])
            if ai is not None:
                if ai in flight_iv_dict:
                    no_ov += flight_iv_dict[ai]
                if ai in break_iv_dict:
                    no_ov.append(break_iv_dict[ai])
            
            model.AddNoOverlap(no_ov)
            can_cover.append(is_cov)
        
        if can_cover:
            model.Add(sum(can_cover) >= 1)

    # Workload optimization objective
    carga_ag = [sum(x[ai][vi] for vi in range(V)) for ai in range(A)]
    for ai in range(A):
        for aj in range(A):
            if ai != aj and activos[ai]['_jornada'] > activos[aj]['_jornada'] + 60:
                model.Add(carga_ag[ai] >= carga_ag[aj])

    # Objective functions based on Mode
    if MODO == 'EQUILIBRADO':
        max_c = model.NewIntVar(0, V, 'mc')
        min_c = model.NewIntVar(0, V, 'nc')
        model.AddMaxEquality(max_c, carga_ag)
        model.AddMinEquality(min_c, carga_ag)
        diff = model.NewIntVar(0, V, 'd')
        model.Add(diff == max_c - min_c)
        model.Minimize(diff)
    elif MODO == 'PROPORCIONAL':
        model.Maximize(sum(x[ai][vi]*activos[ai]['_jornada'] for ai in range(A) for vi in range(V)))

    # Solve model
    solver.parameters.max_time_in_seconds = 60.0
    status = solver.Solve(model)

    # Capture standard output for the 5 Colab Reports!
    stdout_capture = io.StringIO()
    
    if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        # 3. BUILD REVENUE & REPORT OUTPUTS (FROM COLAB OUTPUT SECTIONS)
        asign = {vkey(v): [] for v in VUELOS}
        por_ag = {a['id']: [] for a in activos}
        for ag in AGENTES:
            if ag.get('excluir_embarque') and not ag['excluir']:
                por_ag[ag['id']] = []
        for ai, ag in enumerate(activos):
            for vi, v in enumerate(VUELOS):
                if solver.Value(x[ai][vi]) == 1:
                    asign[vkey(v)].append(ag)
                    por_ag[ag['id']].append(v)

        with contextlib.redirect_stdout(stdout_capture):
            # ─────────────────────────────────────────────────────────────
            # SALIDA 1 — TURNOS DEL PERSONAL
            # ─────────────────────────────────────────────────────────────
            print("═"*72); print("TURNOS DEL PERSONAL"); print("═"*72)
            secciones = [
                ("TURNO MAÑANA · Roles Administrativos - Operativos",
                 [a for a in AGENTES if is_non_boarding_role(a['rol']) and a['_t_ini'] < SPLIT]),
                ("TURNO MAÑANA · Agentes de Pasaje",
                 [a for a in AGENTES if not is_non_boarding_role(a['rol']) and a['_t_ini'] < SPLIT and not a['excluir'] and not a.get('excluir_embarque')]),
                ("TURNO MAÑANA · Solo cobertura de departamento",
                 [a for a in AGENTES if a.get('excluir_embarque') and not a['excluir'] and a['_t_ini'] < SPLIT]),
                ("TURNO MAÑANA · Excluidos hoy",
                 [a for a in AGENTES if a['excluir'] and a['_t_ini'] < SPLIT]),
                ("TURNO TARDE · Roles Administrativos - Operativos",
                 [a for a in AGENTES if is_non_boarding_role(a['rol']) and a['_t_ini'] >= SPLIT]),
                ("TURNO TARDE · Agentes de Pasaje",
                 [a for a in AGENTES if not is_non_boarding_role(a['rol']) and a['_t_ini'] >= SPLIT and not a['excluir'] and not a.get('excluir_embarque')]),
                ("TURNO TARDE · Solo cobertura de departamento",
                 [a for a in AGENTES if a.get('excluir_embarque') and not a['excluir'] and a['_t_ini'] >= SPLIT]),
                ("TURNO TARDE · Excluidos hoy",
                 [a for a in AGENTES if a['excluir'] and a['_t_ini'] >= SPLIT]),
            ]
            for titulo, grupo in secciones:
                if not grupo: continue
                print(f"\n── {titulo}")
                for a in grupo:
                    b2 = f" / {a['bloque2']['inicio']}–{a['bloque2']['fin']}" if a['bloque2'] else ""
                    esp = f" [{'/'.join(a['espec'])}]" if a['espec'] and get_base_role(a['rol']) not in ROLES_OPERATIVOS else ""
                    nota = " ← SOLO COBERTURA" if a.get('excluir_embarque') else (" ← EXCLUIDO" if a['excluir'] else "")
                    print(f"  {a['nombre']:<15} {a['inicio']}–{a['fin']}{b2:<20} {a['rol']}{esp}{nota}")
            n_excluir_emb = sum(1 for a in AGENTES if a.get('excluir_embarque') and not a['excluir'])
            print(f"\n  Total nómina: {len(AGENTES)} · CSA embarcan: {len(activos)} · Solo cobertura: {n_excluir_emb} · Ausentes: {sum(1 for a in AGENTES if a['excluir'])}")
            print(f"  Modo: {MODO}"); print("─"*72)

            # ─────────────────────────────────────────────────────────────
            # SALIDA 2 — PARRILLA
            # ─────────────────────────────────────────────────────────────
            print("\n"+"═"*72); print(f"PARRILLA DE EMBARQUES  [{MODO}]"); print("═"*72)
            print(f"{'Nº':<4}{'DEST':<5}{'VUELO':<10}{'APERTU':<8}{'AGENTES':<30}{'EMB':<7}{'STD'}"); print("─"*72)
            for v in VUELOS_ORD:
                ags = asign[vkey(v)]
                noms = [a['nombre'] for a in ags]
                suf = (' (CH)' if v['charter'] else '') + (' [MAN]' if v['manual'] else '')
                warn = ' ⚠️' if len(noms) < v['agentes_req'] else ''
                if len(noms) == 0:
                    ag_s = "⛔ SIN AGENTES"
                elif len(noms) == 1:
                    ag_s = noms[0]
                elif len(noms) == 2:
                    ag_s = f"{noms[0]} / {noms[1]}"
                else:
                    ag_s = f"{noms[0]} / {noms[1]} / {noms[2]}"
                print(f"{v['num']:<4}{v['destino']:<5}{v['codigo']+suf:<10}{m2t(v['std_min']-180):<8}{ag_s+warn:<30}{m2t(v['emb_inicio']):<7}{v['std']}")

            # ─────────────────────────────────────────────────────────────
            # SALIDA 3 — TOLERANCIAS
            # ─────────────────────────────────────────────────────────────
            print("\n"+"═"*72); print("TOLERANCIAS"); print("═"*72); hay_tol = False
            for ag in activos:
                vag = sorted(por_ag[ag['id']], key=lambda v: v['emb_inicio'])
                disp = ag['_t_ini'] + 30
                for idx, v in enumerate(vag):
                    de = disp - v['emb_inicio']
                    if de > 0:
                        hay_tol = True
                        cat = "Tolerable" if de <= 10 else "No aceptable"
                        icon = "⚠️ " if de <= 10 else "🔴"
                        print(f"{icon} {ndisp(ag)} ({hor_turno_completo(ag)}) — L{v['num']} {v['destino']}: EMB {m2t(v['emb_inicio'])}, entrada {ag['inicio']}+30={m2t(disp)}. ❌ {de} min. {cat}.")
                    if idx == len(vag) - 1:
                        ex = v['std_min'] + 15 - ag['_t_fin']
                        if ex > 0:
                            hay_tol = True
                            cat = "Tolerable" if ex <= 5 else "No aceptable"
                            icon = "⚠️ " if ex <= 5 else "🔴"
                            print(f"{icon} {ndisp(ag)} ({hor_turno_completo(ag)}) — L{v['num']} {v['destino']}: STD {v['std']}+15={m2t(v['std_min']+15)} > {m2t(ag['_t_fin'])}. ❌ {ex} min. {cat}.")
                    if idx < len(vag) - 1:
                        sig = vag[idx+1]
                        gr = sig['emb_inicio'] - v['emb_inicio']
                        gm = gap_minimo(v['zona'], sig['zona'])
                        gt = gap_tolerancia(v['zona'], sig['zona'])
                        zs = "misma zona" if v['zona'] == sig['zona'] else "distinta zona"
                        if gr < gm:
                            hay_tol = True
                            deficit = gm - gr
                            cat = "Tolerable" if gr >= gt else ("Solo si inevitable" if deficit <= 15 else "No aceptable")
                            icon = "⚠️ " if gr >= gt else ("🟡" if deficit <= 15 else "🔴")
                            print(f"{icon} {ndisp(ag)} — L{v['num']} {v['destino']} → L{sig['num']} {sig['destino']}: {m2t(v['emb_inicio'])}→{m2t(sig['emb_inicio'])}: {gr} min [{zs}]. ❌ {deficit} min. {cat}.")
            if not hay_tol:
                print("✅ Sin tolerancias")

            # ─────────────────────────────────────────────────────────────
            # SALIDA 4 — LISTADO POR CANTIDAD DE EMBARQUES
            # ─────────────────────────────────────────────────────────────
            print("\n"+"═"*72); print("LISTADO POR CANTIDAD DE EMBARQUES"); print("═"*72)
            _emb = [a for a in todos_csa if len(por_ag.get(a['id'], [])) > 0]
            _sin = [a for a in todos_csa if len(por_ag.get(a['id'], [])) == 0]
            print(f"👥 Agentes CSA en listado: {len(todos_csa)}   ✈️ Con embarques: {len(_emb)}   💤 Sin embarques: {len(_sin)}")
            for ag in sorted(todos_csa, key=lambda a: (-len(por_ag.get(a['id'], [])), a['_t_ini'], -a['_jornada'], a['nombre'])):
                vag = sorted(por_ag[ag['id']], key=lambda v: v['emb_inicio'])
                n = len(vag)
                comp = {}
                for v in vag:
                    for ag2 in asign[vkey(v)]:
                        if ag2['nombre'] != ag['nombre']:
                            comp[ag2['nombre']] = comp.get(ag2['nombre'], 0) + 1
                comp_s = ', '.join([f"siempre con {c}" if cnt == n and n > 0 else f"con {c} ({cnt}x)" for c, cnt in comp.items()])
                lbl = f"{n} embarque{'s' if n != 1 else ''}"
                print(f"\n{lbl}:")
                if n == 0:
                    print(f"  {ndisp(ag)} ({hor_turno_completo(ag)}) — Sin asignación")
                else:
                    print(f"  {ndisp(ag)} ({hor_turno_completo(ag)}), {comp_s}")
                    for v in vag:
                        print(f"    L{v['num']:>2} {v['destino']} {v['codigo']}  EMB {m2t(v['emb_inicio'])} STD {v['std']}")

            # ─────────────────────────────────────────────────────────────
            # SALIDA 5 — DESCANSOS
            # ─────────────────────────────────────────────────────────────
            print("\n"+"═"*72); print("DESCANSOS"); print("═"*72)
            _oblig_desc = [a for a in todos_csa if a['_jornada'] > 360]
            _recom_desc = [a for a in todos_csa if a['_jornada'] == 360]
            _op_desc = [a for a in AGENTES if get_base_role(a['rol']) in ROLES_OPERATIVOS and get_full_shift_status(a['rol']) != 'SICK' and not a['excluir'] and a['_jornada'] > 360]
            print(f"👥 Resumen: 🔴 CSA >6h (obligatorio): {len(_oblig_desc)}  🟡 CSA =6h (recomendable): {len(_recom_desc)}  🔵 Operativo >6h: {len(_op_desc)}")
            
            def tramos_agente(ag):
                vag = sorted(por_ag[ag['id']], key=lambda v: v['emb_inicio'])
                t = []
                if vag and vag[0]['emb_inicio'] > ag['_t_ini']:
                    t.append((ag['_t_ini'], vag[0]['emb_inicio'], 'inicio jornada', f"L{vag[0]['num']} {vag[0]['destino']}"))
                for i in range(len(vag)-1):
                    v1, v2 = vag[i], vag[i+1]
                    if v2['emb_inicio'] > v1['std_min']:
                        t.append((v1['std_min'], v2['emb_inicio'], f"L{v1['num']} {v1['destino']}", f"L{v2['num']} {v2['destino']}"))
                if vag:
                    u = vag[-1]
                    if ag['_t_fin'] > u['std_min']:
                        t.append((u['std_min'], ag['_t_fin'], f"L{u['num']} {u['destino']}", 'fin jornada'))
                return vag, t

            def imprimir_descanso_csa(ag):
                vag, t = tramos_agente(ag)
                mid = ag['_midpoint']
                exc_emb = ag.get('excluir_embarque', False)
                nota = " · solo cobertura" if exc_emb else ""
                print(f"\n{ndisp(ag)} ({ag['inicio']}–{ag['fin']} / {dur_str(ag['_jornada'])} ){nota} — {len(vag)} embarques")
                if not vag:
                    print("  Todo el turno libre ✅")
                    if exc_emb and ag['espec']:
                        print(f"  Disponible para cobertura [{'/'.join(ag['espec'])}] durante todo su turno")
                    return
                mid_ok = False
                hay_ok_2a = False
                for ts, te, d, h in t:
                    if not mid_ok and ts >= mid:
                        print(f"  ── desde {m2t(mid)} descanso preferible ──────────────")
                        mid_ok = True
                    elif not mid_ok and te > mid:
                        print(f"  ── mitad de jornada: {m2t(mid)} ──────────────────────")
                        mid_ok = True
                    mins = te - ts
                    avail = te - max(ts, mid) if te > mid else 0
                    icon = icono_break(avail) if avail > 0 else ""
                    print(f"  Entre {d} ({m2t(ts)}) y {h} ({m2t(te)}): {dur_str(mins)} {icon}")
                    if avail >= 70:
                        hay_ok_2a = True
                if not mid_ok:
                    print(f"  ── desde {m2t(mid)} descanso preferible ──────────────")
                if ag['_jornada'] > 360 and not hay_ok_2a:
                    max_a = max((min(te, ag['_t_fin']) - max(ts, mid) for ts, te, _, __ in t if te > mid), default=0)
                    print(f"  ⚠️  Sin tramo ≥70 min desde {m2t(mid)} — disponible: {dur_str(max_a)} — PSM")

            def ventanas_libres(ag, desde, hasta):
                vag = sorted(por_ag.get(ag['id'], []), key=lambda v: v['emb_inicio'])
                wins, prev = [], desde
                for v in vag:
                    if v['std_min'] <= desde:
                        prev = max(prev, v['std_min'])
                        continue
                    if v['emb_inicio'] >= hasta:
                        break
                    gs = max(prev, desde)
                    ge = min(v['emb_inicio'], hasta)
                    if ge - gs >= 55:
                        wins.append((gs, ge, ge-gs))
                    prev = max(prev, v['std_min'])
                gs = max(prev, desde)
                if hasta - gs >= 55:
                    wins.append((gs, hasta, hasta-gs))
                return wins

            def descanso_requerido_cobertura(ag):
                if ag['_jornada'] > 360:
                    return 70
                if ag['_jornada'] == 360:
                    return 55
                return 0

            def hay_descanso_disjunto(ag, ex_s, ex_e, req):
                if req <= 0:
                    return True
                for ws, we, wd in ventanas_libres(ag, ag['_midpoint'], ag['_t_fin']):
                    if min(we, ex_s) - ws >= req:
                        return True
                    if we - max(ws, ex_e) >= req:
                        return True
                return False

            def cobertura_dept(dept, op_ag, desde, hasta):
                result = []
                for col in AGENTES:
                    if get_base_role(col['rol']) == dept and col['id'] != op_ag['id'] and not col['excluir']:
                        ol_s = max(desde, col['_t_ini'])
                        ol_e = min(hasta, col['_t_fin'])
                        if ol_e - ol_s >= 55:
                            result.append(('🔵 Colega', ndisp(col), ol_s, ol_e))
                for cov in cobertura_pool:
                    if dept not in cov['espec']:
                        continue
                    ss = max(desde, cov['_t_ini'])
                    se = min(hasta, cov['_t_fin'])
                    if se - ss < 55:
                        continue
                    wins = ventanas_libres(cov, ss, se)
                    if not wins:
                        continue
                    req = descanso_requerido_cobertura(cov)
                    for ws, we, wd in wins:
                        if wd < 55:
                            continue
                        if req == 0 or hay_descanso_disjunto(cov, ws, we, req) or wd >= req + 55:
                            result.append(('🟢 CSA', ndisp(cov), ws, we))
                return result

            # Sort and print breaks
            sort_desc_key = lambda a: (a['_t_ini'], -a['_jornada'], -len(por_ag.get(a['id'], [])), a['nombre'])
            oblig = sorted([a for a in todos_csa if a['_jornada'] > 360], key=sort_desc_key)
            recom = sorted([a for a in todos_csa if a['_jornada'] == 360], key=sort_desc_key)
            
            if oblig:
                print("\n▶ CSA — DESCANSO OBLIGATORIO (jornada >6h)")
                for ag in oblig:
                    imprimir_descanso_csa(ag)
            if recom:
                print("\n▶ CSA — DESCANSO RECOMENDABLE (jornada =6h)")
                for ag in recom:
                    imprimir_descanso_csa(ag)
            
            op_necesitan = sorted([a for a in AGENTES if get_base_role(a['rol']) in ROLES_OPERATIVOS and get_full_shift_status(a['rol']) != 'SICK' and not a['excluir'] and a['_jornada'] > 360], key=lambda a: (a['_t_ini'], -a['_jornada'], a['nombre']))
            if op_necesitan:
                print("\n▶ PERSONAL OPERATIVO — DESCANSO OBLIGATORIO (jornada >6h)")
                for ag in op_necesitan:
                    dept = get_base_role(ag['rol'])
                    mid = ag['_midpoint']
                    print(f"\n{ndisp(ag)} ({ag['inicio']}–{ag['fin']} / {dur_str(ag['_jornada'])})")
                    cob_norm = cobertura_dept(dept, ag, mid, ag['_t_fin'])
                    cob_antes = cobertura_dept(dept, ag, ag['_t_ini'], mid)
                    if cob_norm:
                        print(f"  Descanso desde {m2t(mid)} — cobertura disponible:")
                        seen = set()
                        for tipo, nombre, s, e in sorted(cob_norm, key=lambda x: -(x[3]-x[2]))[:4]:
                            k = f"{nombre}{s}{e}"
                            if k in seen:
                                continue
                            seen.add(k)
                            print(f"    {tipo} {nombre}: {m2t(s)}–{m2t(e)} ({dur_str(e-s)}) {icono_break(e-s)}")
                    elif cob_antes:
                        print(f"  ⚠️  Sin cobertura desde {m2t(mid)} — caso excepcional")
                        print(f"  Descanso antes de la mitad de jornada:")
                        seen = set()
                        for tipo, nombre, s, e in sorted(cob_antes, key=lambda x: -(x[3]-x[2]))[:4]:
                            k = f"{nombre}{s}{e}"
                            if k in seen:
                                continue
                            seen.add(k)
                            print(f"    {tipo} {nombre}: {m2t(s)}–{m2t(e)} ({dur_str(e-s)}) {icono_break(e-s)}")
                    else:
                        print(f"  ⚠️  Sin cobertura disponible para {dept} — comunicar al PSM")

            # Recomendaciones informativas de cobertura por personal operativo SICK.
            sick_operatives = [
                a for a in AGENTES
                if get_base_role(a['rol']) in ROLES_OPERATIVOS
                and get_full_shift_status(a['rol']) == 'SICK'
            ]
            if sick_operatives:
                print("\n▶ COBERTURA RECOMENDADA POR PERSONAL SICK")
                for sick_ag in sick_operatives:
                    dept = get_base_role(sick_ag['rol'])
                    candidates = []
                    for candidate in cobertura_pool:
                        if dept not in candidate['espec'] or candidate['excluir']:
                            continue
                        overlap_start = max(sick_ag['_t_ini'], candidate['_t_ini'])
                        overlap_end = min(sick_ag['_t_fin'], candidate['_t_fin'])
                        if overlap_end - overlap_start < 55:
                            continue
                        unavailable = any(
                            overlap_start < ex_end and overlap_end > ex_start
                            for ex_start, ex_end in candidate.get('_excl_intervals', [])
                        )
                        if unavailable or is_non_boarding_role(candidate['rol']):
                            continue
                        candidates.append(candidate)
                    if candidates:
                        names = ", ".join(ndisp(candidate) for candidate in candidates[:5])
                        print(
                            f"  ⚠️ {sick_ag['nombre']} [{dept}] SICK "
                            f"{sick_ag['inicio']}–{sick_ag['fin']} — candidatos: {names}"
                        )
                    else:
                        print(
                            f"  🔴 {sick_ag['nombre']} [{dept}] SICK "
                            "— sin candidatos de cobertura disponibles"
                        )
                        
            print("\n"+"═"*72); print(f"FIN — AZUL HANDLING · {MODO}"); print("═"*72)

        # Map mapped results for the frontend daily visual schedule table
        results_mapped = {}
        flight_agent_ids = {}
        flight_agent_names = {}
        for vi, v in enumerate(VUELOS):
            ags = asign[vkey(v)]
            flight_key = str(v['id'])
            flight_agent_ids[flight_key] = [agent['id'] for agent in ags]
            flight_agent_names[flight_key] = [agent['nombre'] for agent in ags]
            # Compatibilidad temporal con la parrilla manual antigua.
            if ags:
                results_mapped[flight_key] = ags[0]['id']

        unassigned_flights = [
            v['id'] for v in VUELOS if not flight_agent_ids.get(str(v['id']))
        ]
        workload_by_agent = {}
        for assigned_ids in flight_agent_ids.values():
            for agent_id in assigned_ids:
                workload_by_agent[agent_id] = workload_by_agent.get(agent_id, 0) + 1
        max_workload = max(workload_by_agent.values(), default=0)

        return {
            "success": True,
            "status": "OPTIMAL" if status in (cp_model.OPTIMAL, cp_model.FEASIBLE) else "FEASIBLE",
            "assignments": results_mapped,
            "flight_agent_ids": flight_agent_ids,
            "flight_agent_names": flight_agent_names,
            "unassigned_flights": unassigned_flights,
            "max_workload": max_workload,
            "report_text": stdout_capture.getvalue()
        }
    else:
        return {
            "success": False,
            "status": "INFEASIBLE",
            "message": "No se pudo encontrar un cuadrante viable respetando tus descansos y horarios."
        }

# DEMO Fallback Data (Sábado 20 Junio)
DEMO_AGENTS = [
    { "id": 1, "name": "CARO", "hours": "04:15-14:15", "role": "DSM", "type": "admin" },
    { "id": 2, "name": "DÉBORA", "hours": "02:45-12:45", "role": "PSM", "type": "admin" },
    { "id": 3, "name": "GASTÓN", "hours": "06:45-16:45", "role": "PSM", "type": "admin" },
    { "id": 4, "name": "MOI", "hours": "04:45-14:45", "role": "OPS", "type": "admin" },
    { "id": 5, "name": "BEGO", "hours": "05:00-13:00", "role": "OPS", "type": "admin" },
    { "id": 6, "name": "LIZ", "hours": "03:30-11:30", "role": "TKT", "type": "admin" },
    { "id": 7, "name": "CAROL", "hours": "09:00-16:00", "role": "TKT", "type": "admin" },
    { "id": 8, "name": "CRISTI", "hours": "07:15-15:15", "role": "LL", "type": "admin" },
    { "id": 9, "name": "IRENE", "hours": "02:45-08:10", "role": "CSA", "type": "pasaje" },
    { "id": 10, "name": "STEFANIA", "hours": "02:45-08:45", "role": "CSA", "type": "pasaje" },
    { "id": 11, "name": "MARU", "hours": "03:30-08:00", "role": "CSA", "type": "pasaje" },
    { "id": 12, "name": "ASMA", "hours": "03:30-08:10", "role": "CSA", "type": "pasaje" },
    { "id": 13, "name": "ALEJANDRA", "hours": "03:45-06:00", "role": "CSA", "type": "pasaje" },
    { "id": 14, "name": "ALEJANDRO", "hours": "03:45-06:00", "role": "CSA", "type": "pasaje" },
    { "id": 15, "name": "CATHERINE", "hours": "04:35-07:10", "role": "CSA", "type": "pasaje" },
    { "id": 16, "name": "VICTORIA", "hours": "04:35-07:25", "role": "CSA", "type": "pasaje" },
    { "id": 17, "name": "GUILLE", "hours": "04:35-07:25", "role": "CSA", "type": "pasaje" },
    { "id": 18, "name": "MARTA A.", "hours": "04:35-11:25", "role": "CSA [TKT]", "type": "pasaje" },
    { "id": 19, "name": "NURIA G.", "hours": "04:35-11:25", "role": "CSA [OPS]", "type": "pasaje" },
    { "id": 20, "name": "OMAR", "hours": "04:35-12:20", "role": "CSA", "type": "pasaje" },
    { "id": 21, "name": "MARTA C.", "hours": "04:35-12:20", "role": "CSA [OPS]", "type": "pasaje" },
    { "id": 22, "name": "ANASTASIYA", "hours": "04:50-07:25", "role": "CSA", "type": "pasaje" },
    { "id": 23, "name": "JORGE N.", "hours": "04:55-11:25", "role": "CSA [TKT]", "type": "pasaje" },
    { "id": 24, "name": "LETI", "hours": "04:55-11:25", "role": "CSA", "type": "pasaje" }
]

DEMO_FLIGHTS = [
    { "id": 1, "destination": "MAN", "airline": "FR", "number": "FR3209", "time": "05:45", "agents": "", "pax": 186 },
    { "id": 2, "destination": "EMA", "airline": "FR", "number": "FR4459", "time": "05:45", "agents": "", "pax": 175 },
    { "id": 3, "destination": "NUE", "airline": "FR", "number": "FR5094", "time": "05:45", "agents": "", "pax": 185 },
    { "id": 4, "destination": "BOH", "airline": "FR", "number": "FR5945", "time": "06:00", "agents": "", "pax": 173 },
    { "id": 5, "destination": "BRE", "airline": "FR", "number": "FR9929", "time": "06:05", "agents": "", "pax": 188 },
    { "id": 6, "destination": "LBA", "airline": "FR", "number": "FR2447", "time": "06:15", "agents": "", "pax": 182 },
    { "id": 7, "destination": "MME", "airline": "FR", "number": "FR3374", "time": "06:20", "agents": "", "pax": 179 },
    { "id": 8, "destination": "BUD", "airline": "FR", "number": "FR2274", "time": "06:25", "agents": "", "pax": 181 },
    { "id": 9, "destination": "WMI", "airline": "FR", "number": "FR4059", "time": "06:30", "agents": "", "pax": 176 },
    { "id": 10, "destination": "TNG", "airline": "FR", "number": "FR9587", "time": "06:30", "agents": "", "pax": 186 },
    { "id": 11, "destination": "CRL", "airline": "FR", "number": "FR1915", "time": "06:45", "agents": "", "pax": 192 },
    { "id": 12, "destination": "AAR", "airline": "FR", "number": "FR4695", "time": "06:50", "agents": "", "pax": 174 },
    { "id": 13, "destination": "OPO", "airline": "FR", "number": "FR5046", "time": "06:55", "agents": "", "pax": 201 },
    { "id": 14, "destination": "RAK", "airline": "FR", "number": "FR3909", "time": "07:00", "agents": "", "pax": 187 },
    { "id": 15, "destination": "FMO", "airline": "FR", "number": "FR3368", "time": "07:10", "agents": "", "pax": 189 },
    { "id": 16, "destination": "GOT", "airline": "FR", "number": "FR91",   "time": "07:10", "agents": "", "pax": 194 },
    { "id": 17, "destination": "BER", "airline": "FR", "number": "FR233",  "time": "07:20", "agents": "", "pax": 197 },
    { "id": 18, "destination": "ABZ", "airline": "FR", "number": "FR8007", "time": "07:25", "agents": "", "pax": 190 },
    { "id": 19, "destination": "EIN", "airline": "FR", "number": "FR2575", "time": "07:45", "agents": "", "pax": 189 },
    { "id": 20, "destination": "VLC", "airline": "FR", "number": "FR645",  "time": "08:10", "agents": "", "pax": 189 },
    { "id": 21, "destination": "BCN", "airline": "FR", "number": "FR3081", "time": "08:20", "agents": "", "pax": 187 },
    { "id": 22, "destination": "FCO", "airline": "FR", "number": "FR6139", "time": "09:05", "agents": "", "pax": 190 },
    { "id": 23, "destination": "PRG", "airline": "FR", "number": "FR6658", "time": "09:15", "agents": "", "pax": 191 },
    { "id": 24, "destination": "BLQ", "airline": "FR", "number": "FR8933", "time": "09:25", "agents": "", "pax": 188 },
    { "id": 25, "destination": "VIE", "airline": "FR", "number": "FR703",  "time": "09:55", "agents": "", "pax": 179 },
    { "id": 26, "destination": "ZAG", "airline": "FR", "number": "FR600",  "time": "10:00", "agents": "", "pax": 175 }
]

@app.post("/export-extraction-xlsx")
def export_extraction_xlsx(payload: Dict[str, Any] = Body(...)):
    """Exporta por separado el cuadrante visible de turnos o de vuelos."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        export_type = str(payload.get("type") or "").lower().strip()
        document_date = str(payload.get("date") or "Fecha no detectada").strip()
        if export_type not in {"agents", "flights"}:
            raise HTTPException(status_code=422, detail="Tipo de exportación no válido.")

        wb = openpyxl.Workbook()
        ws = wb.active
        dark_fill = PatternFill("solid", fgColor="1F2937")
        blue_fill = PatternFill("solid", fgColor="2563EB")
        warning_fill = PatternFill("solid", fgColor="FEE2E2")
        body_fill = PatternFill("solid", fgColor="F3F4F6")
        white_font = Font(color="FFFFFF", bold=True)
        title_font = Font(color="FFFFFF", bold=True, size=14)
        thin_gray = Side(style="thin", color="D1D5DB")
        border = Border(bottom=thin_gray)

        if export_type == "agents":
            agents = payload.get("agents") or []
            if not isinstance(agents, list) or not agents:
                raise HTTPException(status_code=400, detail="No hay turnos para descargar.")
            ws.title = "Turnos del personal"
            headers = [
                "Turno",
                "Sección",
                "Agente",
                "Horario",
                "Rol",
                "Observaciones automáticas",
            ]
            title = f"AeroShift — Turnos del Personal — {document_date}"
            rows = []
            for agent in agents:
                errors = agent.get("validation_errors") or []
                if not isinstance(errors, list):
                    errors = [str(errors)]
                rows.append([
                    str(agent.get("shift") or "").capitalize(),
                    "Oficina" if str(agent.get("type") or "").lower() == "admin" else "Pasaje",
                    str(agent.get("name") or ""),
                    str(agent.get("hours") or ""),
                    str(agent.get("role") or agent.get("rol") or ""),
                    " | ".join(str(error) for error in errors if error),
                ])
            filename_prefix = "aeroshift_turnos"
            widths = [14, 14, 24, 31, 25, 65]
        else:
            flights = payload.get("flights") or []
            if not isinstance(flights, list) or not flights:
                raise HTTPException(status_code=400, detail="No hay vuelos para descargar.")
            ws.title = "Parrilla de vuelos"
            headers = [
                "N.º",
                "Destino",
                "Aerolínea",
                "Vuelo",
                "Apertura",
                "Agentes",
                "Embarque",
                "STD",
                "PAX",
            ]
            title = f"AeroShift — Parrilla de Vuelos — {document_date}"

            def move_minutes(clock: str, delta: int) -> str:
                try:
                    hour, minute = map(int, str(clock).split(":"))
                    total = (hour * 60 + minute + delta) % 1440
                    return f"{total // 60:02d}:{total % 60:02d}"
                except Exception:
                    return ""

            rows = []
            for index, flight in enumerate(flights, 1):
                std = str(flight.get("time") or "")
                airline = str(flight.get("airline") or "").upper().strip()
                full_flight_number = str(flight.get("number") or "").upper().strip()
                excel_flight_number = full_flight_number
                if airline and full_flight_number.startswith(airline):
                    suffix = full_flight_number[len(airline):].strip()
                    if suffix:
                        excel_flight_number = suffix
                rows.append([
                    index,
                    str(flight.get("destination") or "?"),
                    airline or "?",
                    excel_flight_number or "?",
                    move_minutes(std, -180) or "?",
                    str(flight.get("agents") or ""),
                    move_minutes(std, -40) or "?",
                    std or "?",
                    flight.get("pax") if flight.get("pax") not in (None, "") else "?",
                ])
            filename_prefix = "aeroshift_vuelos"
            widths = [8, 12, 12, 16, 13, 24, 13, 13, 10]

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.fill = blue_fill
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        for column, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=column, value=header)
            cell.fill = dark_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row_index, row in enumerate(rows, 4):
            for column, value in enumerate(row, 1):
                cell = ws.cell(row=row_index, column=column, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
                cell.fill = body_fill
                if value == "?":
                    cell.font = Font(color="EF4444", bold=True)
            if export_type == "agents" and row[-1]:
                for column in range(1, len(headers) + 1):
                    ws.cell(row=row_index, column=column).fill = warning_fill

        for index, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(index)].width = width
        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{3 + len(rows)}"
        ws.sheet_view.showGridLines = False

        safe_date = re.sub(r"[^0-9A-Za-z_-]+", "-", document_date).strip("-") or "sin-fecha"
        filename = f"{filename_prefix}_{safe_date}.xlsx"
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        headers_response = {
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition",
        }
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers_response,
        )
    except HTTPException:
        raise
    except Exception as error:
        print(f"Error exportando Excel: {error}")
        raise HTTPException(status_code=500, detail="No se pudo generar el archivo Excel.")


@app.post("/extract")
async def extract_data(
    model: Optional[str] = "gpt-5.6-luna",
    document_type: str = Query("all", alias="type"),
    files: List[UploadFile] = File(...),
    authorization: Optional[str] = Header(None),
):
    """Extrae turnos o vuelos mediante visión IA."""
    request_started = time.monotonic()
    document_type = (document_type or "all").lower().strip()
    if document_type not in {"agents", "flights", "all"}:
        raise HTTPException(
            status_code=422,
            detail="El parámetro type debe ser 'agents', 'flights' o 'all'.",
        )

    selected_model = (model or "gpt-5.6-luna").strip()
    is_deepseek = selected_model == "deepseek-v4-flash-vision-exp"
    provider = "deepseek" if is_deepseek else "openai"
    api_key_name = "DEEPSEEK_API_KEY" if is_deepseek else "OPENAI_API_KEY"
    api_key = os.environ.get(api_key_name)
    if not api_key:
        provider_label = "DeepSeek" if is_deepseek else "OpenAI"
        return {
            "success": False,
            "is_real_ai": False,
            "message": (
                f"{provider_label} no está configurado en el servidor. "
                f"Añade {api_key_name} en las variables de entorno de Render."
            ),
            "provider": provider,
            "requested_model": selected_model,
            "date": "Fecha no detectada",
            "agents": [],
            "flights": [],
        }

    try:
        import openpyxl
        import pypdf
        from PIL import Image, ImageEnhance, ImageFilter, ImageOps
        from openai import OpenAI
        
        # La extracción completa debe finalizar antes del límite de tres minutos y medio del navegador.
        # Desactivamos los reintentos internos del SDK para evitar esperas ocultas.
        client_options = {
            "api_key": api_key,
            "timeout": 120.0,
            "max_retries": 0,
        }
        if is_deepseek:
            client_options["base_url"] = "https://api.deepseek.com"
        client = OpenAI(**client_options)

        user_content_blocks = []
        text_payloads = []
        verification_image_blocks = []

        def build_verification_image_block(crop, target_width: int):
            """Amplía y mejora un recorte para la segunda lectura de horarios."""
            gray = ImageOps.grayscale(crop)
            enhanced = ImageOps.autocontrast(gray, cutoff=1)
            enhanced = ImageEnhance.Contrast(enhanced).enhance(1.25)
            enhanced = ImageEnhance.Sharpness(enhanced).enhance(2.0)

            width, height = enhanced.size
            scale = min(
                4.0,
                target_width / max(width, 1),
                3000 / max(height, 1),
            )
            if scale > 1.0:
                enhanced = enhanced.resize(
                    (int(width * scale), int(height * scale)),
                    Image.Resampling.LANCZOS,
                )
            enhanced = enhanced.filter(ImageFilter.SHARPEN).convert("RGB")

            buffer = io.BytesIO()
            enhanced.save(buffer, format="JPEG", quality=92, optimize=True)
            encoded = base64.b64encode(buffer.getvalue()).decode("utf-8")
            image_payload = {"url": f"data:image/jpeg;base64,{encoded}"}
            if not is_deepseek:
                image_payload["detail"] = "high"
            return {
                "type": "image_url",
                "image_url": image_payload,
            }

        for file in files:
            filename_lower = file.filename.lower()
            content = await file.read()
            
            # --- CASO 1: EXCEL (.xlsx, .xls) ---
            if filename_lower.endswith(('.xlsx', '.xls')):
                try:
                    excel_file = io.BytesIO(content)
                    wb = openpyxl.load_workbook(excel_file, data_only=True)
                    
                    sheet_texts = []
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        rows_str = []
                        for row in ws.iter_rows(values_only=True):
                            row_str = " | ".join(str(val).strip() if val is not None else "" for val in row)
                            if row_str.replace(" |", "").strip():
                                rows_str.append(row_str)
                        
                        sheet_texts.append(f"### Hoja: {sheet_name}\n" + "\n".join(rows_str))
                    
                    excel_text = "\n\n".join(sheet_texts)
                    text_payloads.append(f"--- CONTENIDO EXCEL ({file.filename}) ---\n{excel_text}\n")
                except Exception as excel_err:
                    print(f"Error leyendo Excel {file.filename}: {excel_err}")
            
            # --- CASO 2: PDF (.pdf) ---
            elif filename_lower.endswith('.pdf'):
                try:
                    pdf_file = io.BytesIO(content)
                    pdf_reader = pypdf.PdfReader(pdf_file)
                    pdf_text_list = []
                    for page_idx, page in enumerate(pdf_reader.pages):
                        txt = page.extract_text()
                        if txt:
                            pdf_text_list.append(f"--- Página {page_idx + 1} ---\n{txt}")
                    
                    pdf_text = "\n\n".join(pdf_text_list)
                    text_payloads.append(f"--- CONTENIDO PDF ({file.filename}) ---\n{pdf_text}\n")
                except Exception as pdf_err:
                    print(f"Error leyendo PDF {file.filename}: {pdf_err}")
            
            # --- CASO 3: IMÁGENES ---
            else:
                try:
                    base64_image = base64.b64encode(content).decode('utf-8')
                    mime_type = "image/png"
                    if filename_lower.endswith(('.jpg', '.jpeg')):
                        mime_type = "image/jpeg"
                    elif filename_lower.endswith('.gif'):
                        mime_type = "image/gif"
                    elif filename_lower.endswith('.webp'):
                        mime_type = "image/webp"
                    
                    image_payload = {
                        "url": f"data:{mime_type};base64,{base64_image}"
                    }
                    # OpenAI admite control explícito de detalle en Chat Completions.
                    # DeepSeek Vision recibe la imagen sin ese campo adicional.
                    if not is_deepseek:
                        image_payload["detail"] = (
                            "high" if document_type == "agents" else "auto"
                        )
                    user_content_blocks.append({
                        "type": "image_url",
                        "image_url": image_payload,
                    })

                    # Para la segunda lectura de turnos generamos tres vistas:
                    # tabla completa, mitad de mañana y mitad de tarde.
                    if document_type == "agents":
                        try:
                            source_image = ImageOps.exif_transpose(
                                Image.open(io.BytesIO(content))
                            ).convert("RGB")
                            image_width, image_height = source_image.size
                            table_top = max(0, int(image_height * 0.04))
                            table_bottom = min(image_height, int(image_height * 0.66))
                            table_crop = source_image.crop(
                                (0, table_top, image_width, table_bottom)
                            )
                            overlap = int(image_width * 0.01)
                            middle = image_width // 2
                            morning_crop = source_image.crop(
                                (0, table_top, min(image_width, middle + overlap), table_bottom)
                            )
                            afternoon_crop = source_image.crop(
                                (max(0, middle - overlap), table_top, image_width, table_bottom)
                            )
                            verification_image_blocks.extend([
                                build_verification_image_block(table_crop, 1800),
                                build_verification_image_block(morning_crop, 1500),
                                build_verification_image_block(afternoon_crop, 1500),
                            ])
                        except Exception as crop_error:
                            print(
                                f"No se pudieron generar ampliaciones de {file.filename}: "
                                f"{crop_error}"
                            )
                except Exception as img_err:
                    print(f"Error cargando imagen {file.filename}: {img_err}")

        # La interfaz ya informa de qué botón procede el archivo. No pedimos al
        # modelo que mezcle turnos y vuelos cuando el tipo es conocido.
        if document_type == "agents":
            task_instruction = (
                "Extrae exclusivamente la fecha y los turnos del personal. "
                "Devuelve un solo objeto por persona siguiendo el esquema JSON solicitado."
            )
        elif document_type == "flights":
            task_instruction = (
                "Extrae exclusivamente la fecha y la parrilla de vuelos o embarques "
                "siguiendo el esquema JSON solicitado."
            )
        else:
            task_instruction = (
                "Extrae la fecha, los agentes y los vuelos siguiendo el esquema JSON solicitado."
            )

        user_content_blocks.append({
            "type": "text",
            "text": task_instruction,
        })

        if text_payloads:
            combined_text = "\n".join(text_payloads)
            user_content_blocks.append({
                "type": "text",
                "text": (
                    "Aquí tienes el contenido de texto extraído directamente del archivo Excel/PDF cargado.\n"
                    f"Tarea: {task_instruction}\n\n"
                    f"{combined_text}"
                ),
            })

        system_prompt = (
            "Eres un sistema especializado en la extracción documental de máxima precisión.\n"
            "Tu única función es extraer literalmente todo el contenido visible de la imagen.\n"
            "La imagen puede contener tablas, texto impreso, texto manuscrito, sellos, firmas, columnas, encabezados y anotaciones.\n"
            "Debes analizar la imagen varias veces antes de generar la respuesta.\n\n"
            "Normas obligatorias:\n"
            "- No resumas el contenido.\n"
            "- No traduzcas el texto.\n"
            "- No interpretes el significado.\n"
            "- No corrijas errores ortográficos.\n"
            "- No inventes información.\n"
            "- No completes información que no sea visible.\n"
            "- Conserva exactamente las mayúsculas, las minúsculas, los acentos, la puntuación y el orden del documento.\n"
            "- Si existe alguna duda sobre un valor, indícalo con el prefijo «INCERTIDUMBRE:».\n"
            "- Comprueba varias veces el resultado antes de devolver la respuesta.\n"
            "- Comprueba que todas las filas de la tabla se hayan extraído correctamente.\n"
            "- Comprueba que no existan nombres sin horarios.\n"
            "- Comprueba que no existan horarios sin nombres.\n"
            "- Comprueba que el número de elementos extraídos coincida con el número de elementos visibles.\n"
            "- Si la imagen contiene varias tablas, extrae cada una por separado.\n"
            "- Si la imagen contiene una estructura de filas y columnas, respétala.\n"
            "- Si una línea aparece cortada, extrae únicamente la parte visible.\n"
            "- Si el texto está parcialmente oculto, indica que es ilegible.\n"
            "- Nunca omitas información.\n\n"
            "Proceso interno obligatorio:\n"
            "1. Analiza la imagen completa.\n"
            "2. Analiza la parte superior.\n"
            "3. Analiza la parte central.\n"
            "4. Analiza la parte inferior.\n"
            "5. Analiza la columna izquierda.\n"
            "6. Analiza la columna derecha.\n"
            "7. Compara todos los resultados.\n"
            "8. Corrige posibles errores.\n"
            "9. Genera la respuesta final.\n\n"
            "Devuelve exclusivamente un objeto JSON.\n"
            "La estructura debe ser la siguiente:\n"
            "{\n"
            "  \"fecha\": \"DOMINGO 21 JUNIO\",\n"
            "  \"titulo\": \"AZUL HANDLING\",\n"
            "  \"observaciones\": \"\",\n"
            "  \"manana\": [\n"
            "    {\"tipo_elemento\": \"agente\", \"nombre\": \"CARO\", \"horario\": \"04:15-14:15\", \"rol\": \"DSM\"}\n"
            "  ],\n"
            "  \"tarde\": [\n"
            "    {\"tipo_elemento\": \"agente\", \"nombre\": \"SERGIO\", \"horario\": \"14:00-00:00\", \"rol\": \"DSM\"}\n"
            "  ]\n"
            "}\n\n"
            "### REGLAS DE ESTRUCTURA INTERNA PARA LOS ELEMENTOS EN \"manana\" y \"tarde\":\n"
            "1. Para los Turnos de Personal (Horarios):\n"
            "- A LA MANO IZQUIERDA de la hoja están los turnos de la MAÑANA (añade a la lista \"manana\"). A LA MANO DERECHA de la TARDE (añade a la lista \"tarde\").\n"
            "- Bloque superior de la columna: turnos administrativos-operativos con el rol en paréntesis, ej: 'AGENTE1 (DSM)' se extrae con nombre 'AGENTE1' y rol 'DSM'. El bloque superior de cada columna tiene exactamente 5 filas impresas (ej: DSM, PSM, OPS, TKT, LL). La quinta fila siempre pertenece al departamento de Llegadas (LL) (ej: 'AGENTE2 (LL) // AGENTE3 (LL)'). Asegúrate de extraer SIEMPRE a estas personas de la 5ª fila como de rol 'LL' y de tipo 'admin' (¡nunca los desplaces al bloque inferior de agentes de pasaje!).\n"
            "- Bloque inferior de la columna: turnos de Pasaje (CSA). Aquí el rol 'CSA' NO aparece escrito (aparece vacío, solo sale el nombre, ej: 'AGENTE4'). Debes extraerlos con rol 'CSA' de forma automática, salvo que lleven restricciones entre paréntesis (ej: 'AGENTE5 (SOMBRA TKT)', que se extrae con rol 'TKT').\n"
            "- Si en una fila hay dos personas separadas por '/' (ej: 'AGENTE6 / AGENTE7 (PSM)' con horarios '02:45-12:45 / 06:45-16:45'), debes crear dos objetos individuales en la lista correspondientemente:\n"
            "  * Uno con nombre 'AGENTE6', horario '02:45-12:45', rol 'PSM'.\n"
            "  * Otro con nombre 'AGENTE7', horario '06:45-16:45', rol 'PSM'.\n"
            "- ¡¡¡REGLA PARA ROLES MIXTOS / RESTRICCIONES HORARIAS EN EL BLOQUE INFERIOR (PASAJE/CSA)!!!:\n"
            "  * Si un agente de pasaje (CSA) tiene una restricción horaria o rol mixto anotado al lado de su nombre entre paréntesis (ej: 'EVA (11:15-15:00 TKD)' o 'EVA (11:15-15:00 TKT)'), debes extraer su rol exactamente como 'CSA (11:15-15:00 TKT)' de forma que la restricción quede incorporada en el campo 'rol' de su JSON, y su nombre sea simplemente 'EVA'. Nunca unas la restricción al nombre del agente (el nombre debe quedar limpio, ej: 'EVA').\n"
            "- ¡¡¡REGLA DE ORO PARA TURNOS PARTIDOS COMPARTIDOS EN EL BLOQUE INFERIOR (PASAJE) (ej: 'VICTORIA, GUILLE' con '12:30-15:35 // 19:45-21:45')!!!:\n"
            "  * Si varios agentes de pasaje aparecen en la misma fila separados por comas o barras (ej: 'VICTORIA, GUILLE') y tienen un horario con doble barra (ej: '12:30-15:35 // 19:45-21:45'), esto representa un Turno Partido COMPLETO que realizan AMBOS agentes. NO es que uno haga la mañana y el otro la noche. Ambos realizan el horario completo con los dos tramos.\n"
            "  * Por lo tanto, debes extraer cada agente por separado en la lista, pero conservando obligatoriamente el horario completo de los dos tramos para cada uno de ellos (ej: objeto para 'VICTORIA' con horario '12:30-15:35 / 19:45-21:45' y objeto para 'GUILLE' con horario '12:30-15:35 / 19:45-21:45'). Nunca recortes el segundo tramo ni asumas que solo trabaja el primero.\n"
            "\n"
            "2. Para la Parrilla de Vuelos / Embarques:\n"
            "Cada vuelo debe extraerse en este formato dentro de \"manana\" (si es temprano) o \"tarde\" (si es tarde):\n"
            "{\n"
            "  \"tipo_elemento\": \"vuelo\",\n"
            "  \"destino\": \"SNN\",\n"
            "  \"linea_aerea\": \"FR\",\n"
            "  \"numero_vuelo\": \"FR2849\",\n"
            "  \"std\": \"05:45\",\n"
            "  \"pax\": 189\n"
            "}\n"
            "- ¡¡¡EXTRACCIÓN DE PASAJEROS (pax)!!!: Busca en la tabla de vuelos la columna titulada 'WEBS' (que representa la cantidad aproximada de pasajeros facturados) y extrae obligatoriamente ese número como entero en el campo 'pax' (ej: si la columna 'WEBS' dice '189', extrae el número entero 189). Si la columna está vacía, no la encuentras o es ilegible, devuelve null y no inventes ningún valor.\n"
            "- ¡¡¡MUY IMPORTANTE PARA LA HORA DEL VUELO (std)!!!: Extrae estrictamente de la columna o celda 'STD' (ej: '5:45' -> '05:45'). ¡NUNCA extraigas la hora de la columna 'APERTU' (ej: '2:45') ni 'CIERRE' (ej: '5:05') como 'std' del vuelo! Si la columna 'APERTU' dice '2:45' y 'STD' dice '5:45', el std que debes extraer es '05:45'.\n"
            "- ¡¡¡REGLA PARA PÁGINAS DE CONTINUACIÓN (SIN CABECERAS) (ej: filas 41 a 80)!!!:\n"
            "  * Si la imagen cargada empieza directamente con números de fila como 41 o superiores y carece de cabecera impresa arriba, las columnas siguen exactamente la misma estructura de la página 1:\n"
            "    - Columna 1: Número de fila (ej: 41)\n"
            "    - Columna 2: Destino (3 letras, ej: BLQ)\n"
            "    - Columna 3: Código de aerolínea (2 letras, ej: FR)\n"
            "    - Columna 4: Número de vuelo (dígitos, ej: 3913)\n"
            "    - Columna 5: Hora de APERTURA (ej: 11:45)\n"
            "    - Columna 6: Hora de EMBARQUE (suele estar vacía)\n"
            "    - Columna 7: Hora de CIERRE (ej: 14:05)\n"
            "    - Columna 8: Hora de salida STD (ej: 14:45)\n"
            "  * Para el campo 'std' (salida), debes extraer obligatoriamente el valor de la Columna 8 (STD), que en la fila 41 es '14:45' (formato HH:MM: '14:45'), y NUNCA de la Columna 5 (APERTU) que es '11:45'. De igual forma para todas las demás filas de continuación.\n"
            "- 'date' (fecha) debe ser la fecha del documento. PRIORIZA siempre leer cualquier fecha manuscrita con bolígrafo de cualquier tinta o color (azul, negro, rojo, etc.) que aparezca anotada a mano en el papel (ej: '24/06/26' o '21/06/26'). Si no la hay, lee la fecha impresa en la cabecera (ej: 'DOMINGO 21 JUNIO'). Si no encuentras ninguna, pon 'Fecha no detectada' por defecto (¡bajo ningún concepto te inventes una fecha ficticia!).\n"
            "- Si el documento SOLO contiene vuelos, las listas de agentes deben estar vacías, y viceversa. No inventes datos ficticios."
        )

        if document_type == "agents":
            system_prompt = """
Eres un extractor especializado exclusivamente en turnos de personal de handling aeroportuario.
Devuelve únicamente JSON válido, sin Markdown ni explicaciones.

OBJETIVO
Extrae la fecha y todas las personas visibles. Cada persona debe ocupar exactamente un objeto independiente. Nunca coloques dos nombres en el mismo objeto.

ESQUEMA EXACTO
{
  "fecha": "25/6/26",
  "agentes": [
    {
      "nombre": "MARINA",
      "horario": "04:15-14:15",
      "rol": "DSM",
      "seccion": "oficina",
      "turno": "mañana",
      "personas_en_fila": 1
    }
  ],
  "incertidumbres": []
}

CAMPOS OBLIGATORIOS DE CADA AGENTE
- nombre: una sola persona, sin rol ni restricción entre paréntesis.
- horario: un tramo "HH:MM-HH:MM" o un turno partido "HH:MM-HH:MM / HH:MM-HH:MM".
- rol: rol final de esa persona.
- seccion: exclusivamente "oficina" o "pasaje".
- turno: exclusivamente "mañana" o "tarde".
- personas_en_fila: número total de personas visibles en la fila fuente original. Repite el mismo número en todos los objetos procedentes de esa fila.

ESTRUCTURA DEL DOCUMENTO
1. La columna izquierda corresponde a "mañana" y la derecha a "tarde".
2. En cada columna, el bloque superior, antes de la línea horizontal separadora, es "oficina". Habitualmente contiene cinco filas fuente: DSM, PSM, OPS, TKT/TKD y LL. Todas las personas de estas filas tienen seccion="oficina".
3. El bloque inferior, después de la separadora, es "pasaje". Todas sus personas tienen seccion="pasaje", incluso si una persona lleva un rol explícito como PSM, OPS, TKT o LL.
4. La sección depende de la posición en el documento, no del rol.

UNA PERSONA POR OBJETO
- Si una fila contiene varios nombres, crea un objeto separado para cada nombre.
- Normalmente dos palabras completas representan personas distintas, pero existen nombres compuestos protegidos que pertenecen a UNA sola persona: "MARÍA JOSÉ", "JOSÉ MARÍA" y "JUAN CARLOS", con o sin acentos.
- Dos palabras completas no protegidas, como "TRINI KAREN", representan dos personas distintas aunque el separador sea tenue o se haya perdido.
- Dentro de la columna PERSONAL, cualquiera de estos símbolos separa personas: coma ",", más "+", guion "-", barra "/", doble barra "//" y punto y coma ";".
- Una inicial o abreviatura con punto, ª u º seguida de un nombre completo pertenece a UNA sola persona: "M. JOSÉ", "J. CARLOS", "Mª CARMEN", "M.ª CARMEN", "Mº CARMEN", "M.º CARMEN" y "MA. CARMEN".
- Conserva exactamente el símbolo visible ª u º y conserva el espacio: escribe "M. JOSÉ", nunca "M.JOSÉ"; escribe "Mª CARMEN", nunca "MªCARMEN".
- Un identificador individual puede contener una abreviatura posterior de hasta CUATRO letras, siempre separada por un espacio: "SARA P.", "MARÍA GAR", "PAULA DLR" y "SANDRA CAST.". Nunca unas físicamente las letras.
- Un punto al final de una abreviatura posterior puede cerrar ese identificador: "SANDRA CAST. ANASTASIYA" representa SANDRA CAST. y ANASTASIYA. El punto de una inicial anterior, como "M. JOSÉ", no separa.
- Cuenta las personas de la fila antes de separarlas y guarda ese total en personas_en_fila para cada objeto resultante.
- El campo nombre final nunca puede contener separadores de personas.
- Ejemplo de pasaje: "TRINI, KAREN" con "12:50-16:25 / 19:00-22:30" produce dos objetos, TRINI y KAREN, ambos con el turno partido completo y personas_en_fila=2.
- Ejemplo de inicial abreviada: "CANDELARIA, M. JOSÉ" produce solo dos objetos, CANDELARIA y M. JOSÉ; nunca tres objetos.
- Ejemplo de inicial abreviada: "J. CARLOS, VITO" produce solo dos objetos, J. CARLOS y VITO.
- Ejemplo de nombre compuesto: "SOPHIE, MIRIAM, JOSE MARIA" produce tres objetos: SOPHIE, MIRIAM y JOSE MARIA.
- Ejemplo de nombre compuesto: "JUAN CARLOS / CRISTI" produce dos objetos: JUAN CARLOS y CRISTI.
- Un rol escrito junto a una persona afecta solo a esa persona, no a las demás de la fila.
Ejemplo fuente: "SUSANA, AMINA, EMI (TKD)" con "09:00-16:00".
Resultado: SUSANA=CSA, AMINA=CSA y EMI=TKT; tres objetos separados con el mismo horario.

ASIGNACIÓN DE HORARIOS
- Oficina con UNA sola persona en la fila: puede tener un horario continuo o un turno partido de dos tramos. En ambos casos usa personas_en_fila=1.
- Ejemplo válido de oficina: "07:15-09:00 / 13:00-15:15" con "SANDRA M (LL)" produce un único objeto para SANDRA M, conserva los dos tramos completos y usa personas_en_fila=1.
- Oficina con VARIAS personas en la fila y varios horarios: empareja estrictamente por posición. Primer horario con primera persona, segundo horario con segunda persona. Cada objeto debe indicar el total real en personas_en_fila.
- Ejemplo de oficina con dos personas: "15:00-18:30 / 18:15-02:15" con "PAULA N / MELODIA (LL)" significa PAULA N=15:00-18:30 y MELODIA=18:15-02:15, ambas LL y personas_en_fila=2. Nunca copies los dos horarios a una misma persona.
- Oficina con varias personas y un único horario: repite el horario para cada persona e indica el total en personas_en_fila.
- Pasaje: si varios nombres comparten una fila con un solo horario, repite ese horario para cada persona.
- Pasaje: si el horario tiene dos tramos separados por "/" o "//", es un turno partido completo. Repite los dos tramos completos para cada persona de esa fila.
- Normaliza el separador del turno partido como " / ", sin perder ningún tramo.
- No intercambies horas de filas contiguas y no inventes horarios predeterminados.
- Lee dos veces cada horario, dígito por dígito, antes de asociarlo al nombre.
- En filas con varios horarios, verifica por separado entrada y salida de cada persona; no copies ni completes un horario basándote en la fila vecina.
- Una salida con hora de reloj menor que la entrada significa que el turno termina al día siguiente.
- Ninguna jornada puede superar 10 horas reales de trabajo.
- En turnos partidos, suma únicamente la duración de los tramos trabajados; el descanso intermedio no forma parte de la jornada.
- Si tu primera lectura produce más de 10 horas, vuelve a inspeccionar la celda porque la lectura es incorrecta.

ROLES
- En oficina, el rol administrativo-operativo indicado en una fila se aplica a TODAS las personas de esa fila, aunque el paréntesis solo aparezca junto al último nombre.
- Ejemplo de oficina: "DÉBORA / GASTÓN (PSM)" produce DÉBORA=PSM y GASTÓN=PSM.
- Ejemplo de oficina: "MOI / BEGO (OPS)" produce MOI=OPS y BEGO=OPS.
- Ninguna persona del bloque de oficina puede tener rol CSA.
- En pasaje, usa CSA cuando no haya un rol o restricción explícitos junto a esa persona.
- En pasaje, un rol escrito junto a una persona afecta solamente a esa persona, no a toda la fila.
- Si una persona de pasaje lleva PSM, OPS, TKT o LL entre paréntesis, conserva ese rol para esa persona, pero mantiene seccion="pasaje".
- Convierte silenciosamente TKD a TKT. En el JSON final solo debe aparecer TKT.
- Distingue un rol completo de un rol mixto por la presencia de horas dentro del paréntesis.
- Rol completo sin horas, ejemplo "LAURA G (OPS)": rol="OPS" durante toda la jornada.
- Rol mixto con horas, ejemplo "MÍRIAM (OPS 12:30-17)" y jornada general 10:00-17:00: rol="CSA (12:30-17:00 OPS)".
- Fuera del intervalo mixto la persona conserva su rol base CSA; dentro del intervalo realiza el rol indicado y no está disponible para embarques.
- Reconoce tanto "(OPS 12:30-17)" como "(12:30-17 OPS)".
- Normaliza horas incompletas: "17" se convierte en "17:00".
- Usa siempre el formato canónico "CSA (HH:MM-HH:MM ROL)" para cualquier restricción mixta, incluyendo OPS, TKT, PSM, LL, SICK, CURSO u otros roles visibles.
- Nunca elimines el intervalo de una anotación mixta y nunca conviertas una anotación con horas en un rol completo.
- Deja el nombre limpio, sin el rol ni las horas entre paréntesis.

ESTADOS Y RESTRICCIONES: SICK, NUEVO, CURSO, SOMBRA, SHADOW, SHADOWING Y FAMI
- Un estado nunca cambia por sí solo la sección física de la persona.
- Si una persona del bloque de oficina está SICK, conserva su departamento base y la sección oficina. Ejemplos: rol="TKT (SICK)", rol="OPS (SICK)" o rol="LL (SICK)".
- Si una persona del bloque de pasaje está SICK, usa seccion="pasaje" y rol="CSA (SICK)".
- SICK significa ausencia completa: no recibe ninguna asignación durante toda la jornada.
- Una persona marcada NUEVO o NEW conserva su sección y usa el rol base con estado, por ejemplo rol="CSA (NUEVO)". No recibe asignaciones hasta ser operativa.
- Puede haber listas visibles fuera de la tabla principal, por ejemplo: "CURSO TKD 09:00-15:00" seguido de varios nombres.
- Una lista externa de CURSO, SOMBRA, SHADOW, SHADOWING o FAMI NO es un bloque de oficina y no acredita automáticamente a nadie como TKT, OPS, LL, PSM o DSM.
- Cada nombre de esas listas externas se incluye como un objeto independiente con seccion="pasaje".
- Conserva el horario indicado y normaliza TKD a TKT.
- Usa rol canónico con rol base, intervalo y motivo completo: rol="CSA (09:00-15:00 CURSO TKT)", rol="CSA (09:00-15:00 SOMBRA TKT)" o rol="CSA (04:45-12:45 SHADOWING OPS)".
- Durante CURSO, SOMBRA, SHADOW o SHADOWING la persona no está disponible ni para embarques ni para coberturas.
- FAMI y cualquier rol operativo temporal impiden embarcar durante el intervalo indicado.
- Si un nombre aparece también dentro de la tabla principal, NO lo fusiones: crea otra línea independiente porque son personas distintas con IDs distintos.
- OR-Tools aplicará las restricciones utilizando el ID interno, no el nombre visible.

FECHA
- Prioriza la fecha manuscrita visible, aunque esté fuera de la tabla.
- Si no existe, usa la fecha impresa.
- Si no puede leerse ninguna fecha, devuelve "Fecha no detectada".
- No inventes fechas.

PRECISIÓN Y VALIDACIÓN ANTES DE RESPONDER
- Recorre mañana-oficina, mañana-pasaje, tarde-oficina y tarde-pasaje.
- Cuenta todas las personas visibles, no solo las filas.
- Comprueba que cada persona tenga un único objeto y su horario correcto.
- Comprueba especialmente caracteres parecidos como I/L, F/P, N/M y nombres con inicial final.
- No omitas las filas de oficina.
- No añadas personas que no sean visibles.
- Si un texto es realmente ilegible, usa "ILEGIBLE" en ese campo y añade una descripción a incertidumbres; nunca sustituyas el dato por un valor inventado.
- No incluyas vuelos, tipo_elemento, títulos, observaciones ni campos distintos de los definidos en el esquema.
""".strip()

        elif document_type == "flights":
            system_prompt = """
Eres un extractor especializado exclusivamente en parrillas de vuelos aeroportuarios.
Devuelve únicamente JSON válido, sin Markdown ni explicaciones.

OBJETIVO
Extrae todos los vuelos visibles, en el mismo orden del documento, sin omitir filas y sin incluir personal.

ESQUEMA EXACTO
{
  "fecha": "21/06/26",
  "titulo": "PARRILLA DE VUELOS",
  "observaciones": "",
  "manana": [
    {
      "tipo_elemento": "vuelo",
      "destino": "SNN",
      "linea_aerea": "FR",
      "numero_vuelo": "FR2849",
      "std": "05:45",
      "pax": 189
    }
  ],
  "tarde": []
}

REGLAS OBLIGATORIAS
- Crea exactamente un objeto por cada fila de vuelo visible.
- No inventes, resumas, traduzcas ni corrijas los datos del documento.
- No incluyas agentes ni objetos con tipo_elemento distinto de "vuelo".
- Conserva el orden de las filas y combina todas las páginas cargadas.
- Usa "manana" para STD anterior a 14:00 y "tarde" para STD desde 14:00.

CAMPOS
- destino: código IATA visible, normalmente tres letras.
- linea_aerea: código visible de la aerolínea, normalmente dos caracteres.
- numero_vuelo: combina el código de aerolínea y los dígitos cuando aparezcan separados. Ejemplo: FR + 2849 = FR2849.
- std: hora de salida extraída exclusivamente de la columna STD y normalizada como HH:MM.
- pax: entero extraído de la columna WEBS.

STD — REGLA CRÍTICA
- Extrae siempre la hora de la columna STD.
- Nunca utilices APERTU, APERTURA, EMBARQUE o CIERRE como STD.
- Ejemplo: APERTU=02:45, CIERRE=05:05 y STD=05:45 produce std="05:45".
- Normaliza 5:45 como 05:45 sin modificar los minutos.

PAX / WEBS
- Busca la columna WEBS y usa su valor como pax.
- Si WEBS está vacío, ausente o realmente ilegible, devuelve pax=null.
- Nunca inventes 186 ni ningún otro número de pasajeros.
- No tomes cifras de otras columnas como pasajeros.

DATOS AUSENTES O ILEGIBLES
- Si destino, línea aérea, número de vuelo, STD o PAX no pueden leerse, usa null en ese campo.
- No sustituyas datos ausentes por MAD, FR, FR000, 12:00, 186 ni otros valores plausibles.
- Un dato desconocido debe seguir siendo desconocido para que la interfaz lo marque y permita corregirlo.

PÁGINAS DE CONTINUACIÓN SIN CABECERA
Si una página empieza directamente con filas 41 o superiores, conserva esta estructura:
1. Número de fila.
2. Destino.
3. Código de aerolínea.
4. Dígitos del vuelo.
5. APERTURA.
6. EMBARQUE.
7. CIERRE.
8. STD.
- Para std usa siempre la columna 8, nunca la columna 5.
- Ejemplo fila 41: APERTURA=11:45 y STD=14:45 produce std="14:45".

FECHA
- Prioriza cualquier fecha manuscrita visible, aunque esté fuera de la tabla.
- Si no existe fecha manuscrita, usa la fecha impresa de la cabecera.
- Si no se detecta ninguna, devuelve "Fecha no detectada".
- No inventes una fecha.

COMPROBACIÓN FINAL
- Recorre cada página de arriba abajo.
- Comprueba que el número de objetos coincide con el número de filas visibles.
- Verifica de nuevo cada STD contra su columna.
- Verifica que destino, aerolínea, número de vuelo y PAX proceden de la misma fila.
- Devuelve únicamente las claves definidas en el esquema.
""".strip()

        # Se utiliza exclusivamente el modelo elegido en la web. Si falla, el
        # usuario puede reintentarlo manualmente o seleccionar otro modelo.
        # Evitamos una cascada automática que podía superar los siete minutos.
        models_to_try = [selected_model]

        protected_compound_names = {
            "MARIA JOSE",
            "JOSE MARIA",
            "JUAN CARLOS",
        }

        def normalize_name_key(value: str) -> str:
            normalized = unicodedata.normalize("NFD", str(value or "").upper())
            without_accents = "".join(
                char for char in normalized if unicodedata.category(char) != "Mn"
            )
            return re.sub(r"\s+", " ", without_accents).strip()

        def is_protected_compound(value: str) -> bool:
            return normalize_name_key(value) in protected_compound_names

        def is_abbreviated_given_name(token: str) -> bool:
            """Reconoce M., J., Mª, M.ª, Mº, M.º, MA. y variantes equivalentes."""
            value = str(token or "").upper().strip()
            return bool(re.fullmatch(
                r"(?:[A-ZÁÉÍÓÚÜÑ](?:\.|\.?[ªº])|MA\.)",
                value,
            ))

        def split_name_groups(name: str) -> list[str]:
            """
            Separa nombres completos unidos y conserva iniciales o abreviaturas
            de nombre, así como abreviaturas cortas del apellido, con espacios.
            """
            text = re.sub(r"\s+", " ", str(name or "").strip())
            if not text:
                return []

            explicit_parts = [
                part.strip()
                for part in re.split(r"\s*(?:/{1,2}|,|;|\+|-)\s*", text)
                if part.strip()
            ]

            def group_part(part: str) -> list[str]:
                tokens = part.split(" ")
                if len(tokens) <= 1 or is_protected_compound(part):
                    return [part]

                groups = []
                index = 0
                while index < len(tokens):
                    token = tokens[index]

                    # Tras un nombre ya iniciado, una inicial con punto se trata
                    # como abreviatura posterior (SARA P.), no como nuevo nombre.
                    if groups and is_abbreviated_given_name(token):
                        groups[-1].append(token)
                        index += 1
                        continue

                    # Nombres compuestos protegidos, con o sin acentos.
                    if (
                        index + 1 < len(tokens)
                        and is_protected_compound(f"{token} {tokens[index + 1]}")
                    ):
                        group = [token, tokens[index + 1]]
                        index += 2
                        while index < len(tokens):
                            if len(group) > 2 and group[-1].endswith("."):
                                break
                            next_token = tokens[index]
                            next_letters = re.sub(
                                r"[^A-ZÁÉÍÓÚÜÑ]", "", next_token.upper()
                            )
                            if is_abbreviated_given_name(next_token) or len(next_letters) >= 5:
                                break
                            group.append(next_token)
                            index += 1
                        groups.append(group)
                        continue

                    # Una inicial/abreviatura explícita se une al nombre completo siguiente.
                    if is_abbreviated_given_name(token) and index + 1 < len(tokens):
                        group = [token, tokens[index + 1]]
                        index += 2
                        while index < len(tokens):
                            if len(group) > 2 and group[-1].endswith("."):
                                break
                            next_token = tokens[index]
                            next_letters = re.sub(
                                r"[^A-ZÁÉÍÓÚÜÑ]", "", next_token.upper()
                            )
                            if is_abbreviated_given_name(next_token) or len(next_letters) >= 5:
                                break
                            group.append(next_token)
                            index += 1
                        groups.append(group)
                        continue

                    # Un punto tras una abreviatura posterior cierra la persona anterior.
                    if groups and len(groups[-1]) > 1 and groups[-1][-1].endswith("."):
                        groups.append([token])
                        index += 1
                        continue

                    letters = re.sub(r"[^A-ZÁÉÍÓÚÜÑ]", "", token.upper())
                    # Hasta cuatro letras se consideran abreviatura posterior.
                    if not groups or len(letters) >= 5:
                        groups.append([token])
                    else:
                        groups[-1].append(token)
                    index += 1

                return [" ".join(group).strip() for group in groups if group]

            result = []
            for explicit_part in explicit_parts:
                result.extend(group_part(explicit_part))
            return result

        def merge_split_initial_agents(data: Any) -> None:
            """Fusiona iniciales y nombres compuestos separados en objetos consecutivos."""
            if not isinstance(data, dict) or not isinstance(data.get("agentes"), list):
                return
            agents = data["agentes"]
            merged_agents = []
            index = 0
            comparable_fields = ("horario", "rol", "seccion", "turno")

            def same_context(first: dict, second: dict) -> bool:
                return all(
                    str(first.get(field) or "").strip().upper()
                    == str(second.get(field) or "").strip().upper()
                    for field in comparable_fields
                )

            def context_run_info(position: int) -> tuple[int, int]:
                start = position
                end = position + 1
                while (
                    start > 0
                    and isinstance(agents[start - 1], dict)
                    and isinstance(agents[position], dict)
                    and same_context(agents[start - 1], agents[position])
                ):
                    start -= 1
                while (
                    end < len(agents)
                    and isinstance(agents[end], dict)
                    and isinstance(agents[position], dict)
                    and same_context(agents[position], agents[end])
                ):
                    end += 1
                expected = max(
                    int(item.get("personas_en_fila") or 1)
                    for item in agents[start:end]
                    if isinstance(item, dict)
                )
                return end - start, expected

            while index < len(agents):
                current = agents[index]
                if (
                    isinstance(current, dict)
                    and index + 1 < len(agents)
                    and isinstance(agents[index + 1], dict)
                ):
                    following = agents[index + 1]
                    current_name = str(current.get("nombre") or "").strip()
                    following_name = str(following.get("nombre") or "").strip()
                    same = same_context(current, following)
                    initial_pair = is_abbreviated_given_name(current_name)
                    compound_pair = is_protected_compound(
                        f"{current_name} {following_name}"
                    )
                    run_length, expected_people = context_run_info(index)
                    should_merge_compound = compound_pair and run_length > expected_people

                    if same and following_name and (initial_pair or should_merge_compound):
                        combined = dict(current)
                        combined["nombre"] = f"{current_name} {following_name}"
                        combined["personas_en_fila"] = max(
                            int(current.get("personas_en_fila") or 1),
                            int(following.get("personas_en_fila") or 1),
                        )
                        merged_agents.append(combined)
                        index += 2
                        continue

                merged_agents.append(current)
                index += 1

            data["agentes"] = merged_agents

        def expand_merged_passage_names(data: Any) -> None:
            """Separa automáticamente nombres unidos en pasaje y comparte su horario."""
            if not isinstance(data, dict) or not isinstance(data.get("agentes"), list):
                return
            expanded_agents = []
            for agent in data["agentes"]:
                if not isinstance(agent, dict):
                    expanded_agents.append(agent)
                    continue
                groups = split_name_groups(agent.get("nombre"))
                section = str(agent.get("seccion") or "").lower().strip()
                if section == "pasaje" and len(groups) > 1:
                    row_people = max(
                        len(groups),
                        int(agent.get("personas_en_fila") or 1),
                    )
                    for group_name in groups:
                        separated_agent = dict(agent)
                        separated_agent["nombre"] = group_name
                        separated_agent["personas_en_fila"] = row_people
                        expanded_agents.append(separated_agent)
                else:
                    expanded_agents.append(agent)
            data["agentes"] = expanded_agents

        def validate_turns_json(data: Any) -> list[dict]:
            """
            Valida la estructura y anota problemas recuperables por agente.
            Solo los errores estructurales impiden continuar la extracción.
            """
            if not isinstance(data, dict):
                raise ValueError("La respuesta de turnos no es un objeto JSON.")
            if not isinstance(data.get("fecha"), str):
                raise ValueError("Falta el campo fecha en la respuesta de turnos.")
            agents = data.get("agentes")
            if not isinstance(agents, list) or not agents:
                raise ValueError("La respuesta de turnos no contiene una lista de agentes.")

            allowed_sections = {"oficina", "pasaje"}
            allowed_shifts = {"mañana", "tarde"}
            required_fields = {
                "nombre", "horario", "rol", "seccion", "turno", "personas_en_fila"
            }
            issues = []

            for index, agent in enumerate(agents):
                if not isinstance(agent, dict):
                    raise ValueError(f"El agente {index + 1} no es un objeto JSON.")
                missing = required_fields - set(agent)
                if missing:
                    raise ValueError(
                        f"Al agente {index + 1} le faltan campos: {sorted(missing)}."
                    )

                name = str(agent.get("nombre") or "").strip()
                if not name:
                    raise ValueError(f"El agente {index + 1} no tiene nombre.")
                name_groups = split_name_groups(name)
                try:
                    people_in_source_row = int(agent.get("personas_en_fila"))
                except (TypeError, ValueError):
                    raise ValueError(
                        f"personas_en_fila inválido en el agente {index + 1}."
                    )
                if people_in_source_row < 1 or people_in_source_row > 50:
                    raise ValueError(
                        f"personas_en_fila fuera de rango en el agente {index + 1}."
                    )
                agent["personas_en_fila"] = people_in_source_row

                agent_errors = []
                agent["_validation_errors"] = agent_errors

                def add_issue(field: str, message: str) -> None:
                    agent_errors.append(message)
                    issues.append({
                        "indice": index,
                        "nombre": name,
                        "campo": field,
                        "mensaje": message,
                    })

                schedule = str(agent.get("horario") or "").strip()
                if not schedule:
                    normalized_schedule = "ILEGIBLE"
                    work_minutes = None
                    add_issue("horario", "Horario vacío o ilegible.")
                else:
                    try:
                        normalized_schedule, work_minutes = normalize_and_measure_schedule(schedule)
                    except Exception:
                        normalized_schedule = schedule.upper()
                        work_minutes = None
                        add_issue(
                            "horario",
                            f"Formato de horario pendiente de revisión: {schedule}.",
                        )
                if work_minutes is not None and work_minutes > 600:
                    add_issue(
                        "horario",
                        f"Jornada superior a 10 horas: {normalized_schedule} "
                        f"({work_minutes} minutos).",
                    )
                agent["horario"] = normalized_schedule

                try:
                    role, mixed_interval = normalize_mixed_role(agent.get("rol") or "")
                except Exception:
                    role = str(agent.get("rol") or "ILEGIBLE").upper().strip()
                    mixed_interval = None
                    add_issue("rol", f"Rol pendiente de revisión: {role}.")
                if not role:
                    role = "ILEGIBLE"
                    add_issue("rol", "Rol vacío o ilegible.")
                if (
                    mixed_interval is None
                    and get_full_shift_status(role) is None
                    and ("(" in role or ")" in role)
                ):
                    add_issue(
                        "rol",
                        f"Restricción mixta incompleta o no reconocida: {role}.",
                    )

                section = str(agent.get("seccion") or "").lower().strip()
                if section not in allowed_sections:
                    raise ValueError(f"Sección inválida en el agente {index + 1}.")
                if len(name_groups) > 1:
                    add_issue(
                        "nombre",
                        "Posible unión de varias personas en un solo nombre: "
                        + " / ".join(name_groups)
                        + ".",
                    )
                if section == "oficina" and role.startswith("CSA"): 
                    add_issue("rol", "Un agente de oficina no puede tener rol CSA.")
                if (
                    section == "oficina"
                    and people_in_source_row > 1
                    and " / " in normalized_schedule
                ):
                    add_issue(
                        "horario",
                        "La fila de oficina contiene varias personas; sus horarios deben asignarse individualmente por posición.",
                    )

                if mixed_interval is not None:
                    restriction_start, restriction_end, restricted_role = mixed_interval
                    if section != "pasaje":
                        add_issue(
                            "rol",
                            "Una restricción mixta debe pertenecer a un agente de pasaje.",
                        )
                    if not is_non_boarding_role(restricted_role):
                        add_issue(
                            "rol",
                            f"Rol restringido no reconocido: {restricted_role}.",
                        )
                    try:
                        interval_is_valid = mixed_interval_is_within_schedule(
                            normalized_schedule, restriction_start, restriction_end
                        )
                    except Exception:
                        interval_is_valid = False
                    if not interval_is_valid:
                        add_issue(
                            "rol",
                            f"La restricción {restriction_start}-{restriction_end} queda fuera de la jornada {normalized_schedule}.",
                        )

                shift = str(agent.get("turno") or "").lower().strip()
                if shift not in allowed_shifts:
                    raise ValueError(f"Turno inválido en el agente {index + 1}.")

                agent["rol"] = role
                agent["seccion"] = section
                agent["turno"] = shift

            return issues

        parsed_result = None
        extracted_model_name = ""
        errors = []
        first_extraction_seconds = None
        first_validation_issues = []

        for model_name in models_to_try:
            try:
                first_call_started = time.monotonic()
                print(f"Intentando llamada a {model_name}...")
                params = {
                    "model": model_name,
                    "messages": [
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": user_content_blocks}
                    ],
                    "response_format": {"type": "json_object"},
                }
                if is_deepseek:
                    params["max_tokens"] = 16000
                else:
                    # En OpenAI incluye razonamiento y JSON final.
                    params["max_completion_tokens"] = 16000
                    params["reasoning_effort"] = "high"
                    
                # Realizamos llamada OpenAI
                try:
                    response = client.chat.completions.create(**params)
                except Exception as call_err:
                    err_msg = str(call_err).lower()
                    print(f"Ajuste defensivo para {model_name} por error: {call_err}")
                    adjusted = False

                    # Solo se reintenta si el servidor rechaza inmediatamente
                    # un parámetro concreto. Un timeout o error de red no genera
                    # una segunda llamada oculta.
                    if "response_format" in err_msg or "json mode" in err_msg:
                        if "response_format" in params:
                            del params["response_format"]
                            adjusted = True

                    if "reasoning_effort" in err_msg or "unsupported_parameter" in err_msg:
                        if "reasoning_effort" in params:
                            del params["reasoning_effort"]
                            adjusted = True

                    if "system role" in err_msg or "unsupported role" in err_msg:
                        merged_content = [
                            {"type": "text", "text": f"INSTRUCCIONES DEL SISTEMA:\n{system_prompt}\n\n"}
                        ]
                        if isinstance(user_content_blocks, list):
                            merged_content.extend(user_content_blocks)
                        else:
                            merged_content.append({"type": "text", "text": str(user_content_blocks)})
                        params["messages"] = [
                            {"role": "user", "content": merged_content}
                        ]
                        adjusted = True

                    if not adjusted:
                        raise
                    response = client.chat.completions.create(**params)

                # Procesamiento y diagnóstico del resultado de este modelo.
                choice = response.choices[0]
                finish_reason = getattr(choice, "finish_reason", None)
                usage = getattr(response, "usage", None)
                raw_result = choice.message.content

                if not raw_result:
                    print(
                        f"Respuesta vacía de {model_name}; finish_reason={finish_reason}; "
                        f"usage={usage}"
                    )
                    raise Exception(
                        "El modelo no generó un resultado utilizable dentro del límite permitido."
                    )

                raw_result = raw_result.strip()
                print(
                    f"Respuesta recibida de {model_name}; "
                    f"finish_reason={finish_reason}; "
                    f"primeros 150 caracteres: '{raw_result[:150]}'"
                )

                # Limpieza de markdown
                if raw_result.startswith("```"):
                    raw_result = re.sub(r"^```(?:json)?\n", "", raw_result, flags=re.IGNORECASE)
                    raw_result = re.sub(r"\n```$", "", raw_result)
                    raw_result = raw_result.strip()

                # Intento de parseo JSON para este modelo
                try:
                    parsed_result = json.loads(raw_result)
                except json.JSONDecodeError as json_err:
                    # Intento defensivo usando búsqueda de llaves con regex
                    match_json = re.search(r"\{.*\}", raw_result, re.DOTALL)
                    if match_json:
                        parsed_result = json.loads(match_json.group(0))
                    else:
                        raise json_err

                # Un JSON sintácticamente válido también debe cumplir el esquema
                # del tipo de documento solicitado antes de detener la cascada.
                if document_type == "agents":
                    merge_split_initial_agents(parsed_result)
                    expand_merged_passage_names(parsed_result)
                    first_validation_issues = validate_turns_json(parsed_result)

                first_extraction_seconds = round(
                    time.monotonic() - first_call_started, 2
                )
                extracted_model_name = model_name
                break

            except Exception as model_err:
                parsed_result = None
                print(f"Fallo de llamada o decodificación con {model_name}: {model_err}")
                errors.append(f"{model_name}: {str(model_err)}")
                continue

        if parsed_result is None:
            last_error = errors[-1].split(": ", 1)[-1] if errors else ""
            lower_error = last_error.lower()
            visible_validation_errors = (
                "jornada imposible",
                "tiene dos tramos",
                "restricción mixta",
                "queda fuera de su jornada",
                "no puede tener rol csa",
            )
            if is_deepseek and any(
                marker in lower_error
                for marker in ("insufficient balance", "insufficient_balance", "402")
            ):
                public_error = (
                    "DeepSeek no dispone de saldo suficiente. "
                    "Recarga la cuenta en platform.deepseek.com."
                )
            elif any(marker in lower_error for marker in ("authentication", "invalid api key", "401")):
                public_error = (
                    f"La clave de {provider.capitalize()} no es válida o no está autorizada."
                )
            elif any(marker in lower_error for marker in visible_validation_errors):
                public_error = last_error[:500]
            elif "timeout" in lower_error or "timed out" in lower_error:
                public_error = (
                    "La extracción superó el tiempo máximo permitido. "
                    "Puedes volver a intentarlo o elegir manualmente otro modelo."
                )
            else:
                public_error = (
                    "El modelo seleccionado no pudo generar una extracción válida. "
                    "Puedes revisar la imagen y volver a intentarlo."
                )
            raise Exception(public_error)

        # Segunda lectura automática y especializada exclusivamente en horarios.
        # Si falla, se conserva íntegramente la primera extracción.
        verification_completed = None
        verification_corrections = 0
        verification_warning = ""
        verification_seconds = None

        if document_type == "agents":
            verification_completed = False
            elapsed = time.monotonic() - request_started
            remaining_budget = 195.0 - elapsed

            if remaining_budget < 25.0:
                verification_warning = (
                    "La extracción principal se completó, pero no quedó tiempo "
                    "suficiente para la segunda verificación de horarios."
                )
            else:
                try:
                    extracted_agents = parsed_result.get("agentes", [])
                    verification_rows = [
                        {
                            "indice": index,
                            "nombre": agent["nombre"],
                            "horario_actual": agent["horario"],
                            "seccion": agent["seccion"],
                            "turno": agent["turno"],
                            "personas_en_fila": agent["personas_en_fila"],
                            "alertas": agent.get("_validation_errors", []),
                        }
                        for index, agent in enumerate(extracted_agents)
                    ]

                    verification_system_prompt = """
Eres un verificador visual especializado exclusivamente en horarios laborales.
Recibirás ampliaciones de la misma parrilla y una lista numerada procedente de una primera extracción.
Las imágenes son vistas complementarias del mismo documento: tabla completa, mañana ampliada y tarde ampliada. No representan personas adicionales.

TAREA
- Revisa visualmente, una por una, TODAS las personas de la lista.
- Revisa primero y con máxima atención las personas cuyo campo alertas no esté vacío.
- Una alerta matemática no confirma el valor correcto: vuelve a leer los dígitos en la imagen.
- Compara exclusivamente horario_actual con la celda correspondiente de la imagen.
- Presta especial atención a los dígitos de la hora de salida.
- No cambies nombres, roles, secciones ni turnos.
- Conserva el horario actual salvo que la diferencia sea claramente visible.
- Si varias personas comparten una misma celda, revisa y corrige a cada una de ellas.
- Oficina con personas_en_fila=1 puede tener uno o dos tramos; conserva el turno partido si es visible.
- Oficina con personas_en_fila>1 debe tener los horarios individuales emparejados por posición, no un turno partido copiado a cada persona.
- En pasaje, conserva completos los turnos partidos cuando realmente existan.
- Normaliza como HH:MM-HH:MM o HH:MM-HH:MM / HH:MM-HH:MM.

Devuelve exclusivamente este JSON:
{
  "total_revisados": 0,
  "correcciones": [
    {
      "indice": 0,
      "nombre": "BEGO",
      "horario_actual": "05:00-15:00",
      "horario_corregido": "05:00-13:00",
      "confianza": "alta"
    }
  ]
}

REGLAS
- total_revisados debe coincidir con el número total de personas recibidas.
- Devuelve únicamente discrepancias claras.
- Usa confianza="alta" solo cuando los dígitos sean claramente legibles.
- Si no hay discrepancias, devuelve correcciones=[] con el total revisado.
""".strip()

                    if verification_image_blocks:
                        # Orden por documento: tabla completa, mañana ampliada y tarde ampliada.
                        verification_content = list(verification_image_blocks)
                    else:
                        # Si Pillow no pudo crear recortes, reutilizamos la imagen original.
                        verification_content = [
                            block
                            for block in user_content_blocks
                            if isinstance(block, dict) and block.get("type") == "image_url"
                        ]
                    verification_content.append({
                        "type": "text",
                        "text": (
                            "Lista de la primera extracción:\n"
                            + json.dumps(verification_rows, ensure_ascii=False)
                        ),
                    })

                    verification_timeout = min(60.0, max(15.0, remaining_budget - 5.0))
                    verification_client = client.with_options(
                        timeout=verification_timeout,
                        max_retries=0,
                    )
                    verification_call_started = time.monotonic()
                    verification_params = {
                        "model": selected_model,
                        "messages": [
                            {"role": "system", "content": verification_system_prompt},
                            {"role": "user", "content": verification_content},
                        ],
                        "response_format": {"type": "json_object"},
                    }
                    if is_deepseek:
                        verification_params["max_tokens"] = 6000
                    else:
                        verification_params["max_completion_tokens"] = 6000
                        verification_params["reasoning_effort"] = "high"
                    verification_response = verification_client.chat.completions.create(
                        **verification_params
                    )
                    verification_raw = verification_response.choices[0].message.content
                    if not verification_raw:
                        raise ValueError("La segunda verificación devolvió contenido vacío.")
                    verification_raw = verification_raw.strip()
                    if verification_raw.startswith("```"):
                        verification_raw = re.sub(
                            r"^```(?:json)?\n", "", verification_raw, flags=re.IGNORECASE
                        )
                        verification_raw = re.sub(r"\n```$", "", verification_raw).strip()
                    verification_result = json.loads(verification_raw)

                    total_reviewed = verification_result.get("total_revisados")
                    corrections = verification_result.get("correcciones")
                    if total_reviewed != len(extracted_agents):
                        raise ValueError(
                            "La segunda verificación no confirmó todas las personas."
                        )
                    if not isinstance(corrections, list):
                        raise ValueError("La segunda verificación no devolvió correcciones válidas.")

                    used_indexes = set()
                    for correction in corrections:
                        try:
                            if not isinstance(correction, dict):
                                continue
                            index = int(correction.get("indice"))
                            if index in used_indexes or not 0 <= index < len(extracted_agents):
                                continue
                            if str(correction.get("confianza") or "").lower().strip() != "alta":
                                continue

                            agent = extracted_agents[index]
                            expected_name = str(agent.get("nombre") or "").upper().strip()
                            correction_name = str(correction.get("nombre") or "").upper().strip()
                            if correction_name != expected_name:
                                continue
                            current_schedule = str(agent.get("horario") or "").strip()
                            if str(correction.get("horario_actual") or "").strip() != current_schedule:
                                continue

                            corrected_schedule, corrected_minutes = normalize_and_measure_schedule(
                                correction.get("horario_corregido")
                            )
                            if corrected_schedule == "ILEGIBLE":
                                continue
                            if corrected_minutes is None or corrected_minutes > 600:
                                continue
                            if (
                                agent.get("seccion") == "oficina"
                                and int(agent.get("personas_en_fila") or 1) > 1
                                and " / " in corrected_schedule
                            ):
                                continue
                            if corrected_schedule == current_schedule:
                                continue

                            agent["horario"] = corrected_schedule
                            used_indexes.add(index)
                            verification_corrections += 1
                        except Exception as correction_error:
                            print(f"Corrección de horario descartada: {correction_error}")
                            continue

                    verification_seconds = round(
                        time.monotonic() - verification_call_started, 2
                    )
                    verification_completed = True
                    print(
                        "Segunda verificación completada: "
                        f"{verification_corrections} correcciones aplicadas."
                    )

                except Exception as verification_error:
                    if "verification_call_started" in locals():
                        verification_seconds = round(
                            time.monotonic() - verification_call_started, 2
                        )
                    print(
                        "La segunda verificación de horarios no se completó: "
                        f"{verification_error}"
                    )
                    verification_warning = (
                        "La extracción principal se completó, pero no pudo finalizarse "
                        "la segunda verificación de horarios. Revisa los horarios antes de importar."
                    )

        final_validation_issues = []
        validation_warning = ""
        if document_type == "agents":
            # Se vuelve a validar después de aplicar las correcciones de la segunda lectura.
            final_validation_issues = validate_turns_json(parsed_result)
            if final_validation_issues:
                validation_warning = (
                    f"Quedan {len(final_validation_issues)} incidencias pendientes. "
                    "Corrige las filas marcadas antes de validar e importar."
                )

        extracted_date = parsed_result.get("fecha") or "Fecha no detectada"
        if not extracted_date or extracted_date.strip() == "":
            extracted_date = "Fecha no detectada"
            
        if document_type == "agents":
            # El nuevo esquema ya contiene exactamente una persona por objeto.
            # Se adapta aquí al formato interno existente sin volver a dividir nombres.
            raw_manana = []
            raw_tarde = []
            for agent in parsed_result.get("agentes", []):
                internal_item = {
                    "tipo_elemento": "agente",
                    "nombre": agent["nombre"],
                    "horario": agent["horario"],
                    "rol": agent["rol"],
                    "seccion": agent["seccion"],
                    "source_row_people": agent["personas_en_fila"],
                    "validation_errors": agent.get("_validation_errors", []),
                }
                turno = str(agent["turno"]).lower().strip()
                if turno == "mañana":
                    raw_manana.append(internal_item)
                else:
                    raw_tarde.append(internal_item)
        else:
            # Compatibilidad temporal para vuelos y cargas antiguas.
            raw_manana = parsed_result.get("manana") or []
            raw_tarde = parsed_result.get("tarde") or []

        formatted_agents = []
        formatted_flights = []
        
        def process_items(items, shift_name):
            for idx, item in enumerate(items):
                if not isinstance(item, dict):
                    continue
                tipo = str(item.get("tipo_elemento") or "").lower().strip()
                
                if not tipo:
                    if "horario" in item or "rol" in item:
                        tipo = "agente"
                    elif "std" in item or "numero_vuelo" in item:
                        tipo = "vuelo"
                        
                if tipo == "agente":
                    name_raw = str(item.get("nombre") or "").upper().strip()
                    hours_raw = str(item.get("horario") or "08:00-16:00").strip()
                    role_upper = str(item.get("rol") or "CSA").upper().strip()
                    
                    # Corrección automática de TKD a TKT para evitar errores ortográficos del papel impreso
                    if "TKD" in role_upper:
                        role_upper = role_upper.replace("TKD", "TKT")
                    
                    # Dividimos nombres múltiples separados por '/' o ','
                    names = [n.strip() for n in name_raw.replace(",", "/").split("/") if n.strip()]
                    
                    # CLASIFICACIÓN DETERMINISTA BASADA EN LA SECCIÓN DEL DOCUMENTO (oficina vs pasaje):
                    # - Si la sección es "oficina", pertenece estrictamente al bloque superior ("admin").
                    # - Si la sección es "pasaje", pertenece estrictamente al bloque inferior de agentes de pasaje ("pasaje").
                    # - Esto evita cualquier desalineación por filas divididas y es 100% robusto con todos los modelos.
                    seccion_raw = str(item.get("seccion") or "").lower().strip()
                    if seccion_raw == "oficina":
                        agent_type = "admin"
                    elif seccion_raw == "pasaje":
                        agent_type = "pasaje"
                    else:
                        # Fallback por si la sección está vacía
                        agent_type = "admin"
                        if "CSA" in role_upper or not role_upper:
                            agent_type = "pasaje"
                            if not role_upper:
                                role_upper = "CSA"

                    if len(names) <= 1:
                        # ¡Un solo agente! Conservamos su horario exactamente igual (preservando turnos partidos con '/')
                        formatted_agents.append({
                            "id": len(formatted_agents) + 1,
                            "name": name_raw,
                            "hours": hours_raw,
                            "role": role_upper,
                            "type": agent_type,
                            "shift": shift_name,
                            "source_row_people": int(item.get("source_row_people") or 1),
                            "validation_errors": item.get("validation_errors", []),
                        })
                    else:
                        # Múltiples agentes en la misma fila (como AgenteA, AgenteB o AgenteC / AgenteD)
                        if agent_type == "pasaje":
                            # Los CSA (pasaje) que comparten fila SIEMPRE comparten el mismo turno partido completo. ¡NO se dividen sus horas!
                            for name_val in names:
                                formatted_agents.append({
                                    "id": len(formatted_agents) + 1,
                                    "name": name_val,
                                    "hours": hours_raw, # Ambos reciben el turno partido completo
                                    "role": role_upper,
                                    "type": agent_type,
                                    "shift": shift_name
                                })
                        else:
                            # Los administrativos (admin) sí que tienen turnos individuales distintos separados por la barra
                            hours_list = [h.strip() for h in hours_raw.replace("//", "/").split("/") if h.strip()]
                            for idx_name, name_val in enumerate(names):
                                hours_val = hours_raw
                                if len(hours_list) == len(names):
                                    hours_val = hours_list[idx_name]
                                elif len(hours_list) > 0:
                                    hours_val = hours_list[min(idx_name, len(hours_list) - 1)]
                                    
                                formatted_agents.append({
                                    "id": len(formatted_agents) + 1,
                                    "name": name_val,
                                    "hours": hours_val,
                                    "role": role_upper,
                                    "type": agent_type,
                                    "shift": shift_name
                                })
                elif tipo == "vuelo":
                    pax_val = item.get("pax")
                    if pax_val is None:
                        pax_val = item.get("webs")
                    try:
                        pax_num = int(pax_val) if pax_val not in (None, "") else None
                    except (TypeError, ValueError):
                        pax_num = None

                    def clean_flight_text(value: Any) -> Optional[str]:
                        if value is None:
                            return None
                        cleaned = str(value).strip()
                        return cleaned.upper() if cleaned else None

                    formatted_flights.append({
                        "id": len(formatted_flights) + 1,
                        "destination": clean_flight_text(item.get("destino")),
                        "airline": clean_flight_text(item.get("linea_aerea")),
                        "number": clean_flight_text(item.get("numero_vuelo")),
                        "time": clean_flight_text(item.get("std")),
                        "agents": "",
                        "pax": pax_num
                    })

        process_items(raw_manana, "mañana")
        process_items(raw_tarde, "tarde")
        total_backend_seconds = round(time.monotonic() - request_started, 2)

        return {
            "success": True,
            "is_real_ai": True,
            "message": f"Extracción completada por {extracted_model_name}.",
            "document_type": document_type,
            "provider": provider,
            "requested_model": selected_model,
            "used_model": extracted_model_name,
            "used_fallback": extracted_model_name != selected_model,
            "verification_completed": verification_completed,
            "verification_corrections": verification_corrections,
            "verification_warning": verification_warning,
            "validation_issues": final_validation_issues,
            "validation_issue_count": len(final_validation_issues),
            "validation_warning": validation_warning,
            "first_extraction_seconds": first_extraction_seconds,
            "verification_seconds": verification_seconds,
            "total_backend_seconds": total_backend_seconds,
            "date": str(extracted_date).strip(),
            "agents": formatted_agents,
            "flights": formatted_flights,
        }

    except Exception as e:
        # El detalle completo queda en los logs de Render, no se expone al navegador.
        tb = traceback.format_exc()
        print(f"Error calling {provider} Vision API: {e}\nTraceback: {tb}")
        return {
            "success": False,
            "is_real_ai": False,
            "message": f"No se pudo completar la extracción: {str(e)}",
            "document_type": document_type,
            "provider": provider,
            "requested_model": selected_model,
            "used_model": None,
            "used_fallback": False,
            "date": "Fecha no detectada",
            "agents": [],
            "flights": [],
        }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
